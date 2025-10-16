from phe import paillier
import random
import numpy as np
import pickle
import os
import math
import time
from functools import partial
from concurrent.futures import ProcessPoolExecutor
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm

# NSGA-II核心实现（加密环境下）
def nsga2_non_dominated_sort_enc(obj_list, pubkey, share0, share1, mu, n, n_sq):
    """
    NSGA-II的非支配排序（支持加密目标值）
    返回：fronts（List[List[int]]）多层非支配解索引
    """
    S = {}  # 支配集合
    n_dom = {}  # 被支配次数
    fronts = [[]]  # 非支配层

    for p in range(len(obj_list)):
        S[p] = []
        n_dom[p] = 0
        for q in range(len(obj_list)):
            if p == q:
                continue
            if dominates_enc(obj_list[p], obj_list[q], pubkey, share0, share1, mu, n, n_sq):
                S[p].append(q)
            elif dominates_enc(obj_list[q], obj_list[p], pubkey, share0, share1, mu, n, n_sq):
                n_dom[p] += 1
        if n_dom[p] == 0:
            fronts[0].append(p)

    i = 0
    while len(fronts[i]) > 0:
        next_front = []
        for p in fronts[i]:
            for q in S[p]:
                n_dom[q] -= 1
                if n_dom[q] == 0:
                    next_front.append(q)
        i += 1
        fronts.append(next_front)

    # 去掉最后一个空层
    if len(fronts[-1]) == 0:
        fronts.pop()

    return fronts


def secure_sort_encrypted_numbers(enc_numbers, pubkey, share0, share1, mu, n, n_sq):
    """
    安全排序加密数列表；返回排序后的索引（升序）
    注意：避免覆盖模数 n
    """
    cnt = len(enc_numbers)                 # 原来这里误写成了 n = len(enc_numbers)
    indices = list(range(cnt))

    for i in range(cnt):
        for j in range(0, cnt - i - 1):
            # 比较 enc_numbers[indices[j]] > enc_numbers[indices[j+1]]
            if secure_greater_than(enc_numbers[indices[j]], enc_numbers[indices[j + 1]],
                                   pubkey, share0, share1, mu, n, n_sq):
                indices[j], indices[j + 1] = indices[j + 1], indices[j]
    return indices


def secure_greater_than(enc_a, enc_b, pubkey, share0, share1, mu, n, n_sq):
    """
    判断 enc_a > enc_b ?
    利用 secure_compare(x, y) ≈ [x < y] 的加密位：
      a > b  等价于  b < a
    """
    enc_bit = secure_compare(enc_b, enc_a, pubkey, share0, share1, mu, n, n_sq)  # [b < a]
    u0 = partial_decrypt(enc_bit, share0, pubkey, n_sq)
    u1 = partial_decrypt(enc_bit, share1, pubkey, n_sq)
    bit = combine_shares(u0, u1, mu, n)
    return bit == 1


# 修改拥挤距离计算函数
def nsga2_crowding_distance_assignment(obj_list, indices, pubkey, share0, share1, mu, n, n_sq):
    """
    NSGA-II的拥挤距离计算（支持加密目标值）
    返回：{个体索引: 拥挤距离}
    """
    n_obj = 2  # 成本和质量两个目标
    distances = {i: 0.0 for i in indices}

    for m in range(n_obj):
        # 提取当前目标的加密值
        enc_values = [obj_list[i][1 + m] for i in indices]

        # 使用安全排序获取索引
        sorted_indices = secure_sort_encrypted_numbers(enc_values, pubkey, share0, share1, mu, n, n_sq)
        # 映射回原始索引
        sorted_indices = [indices[i] for i in sorted_indices]

        # 边界点的拥挤距离设为无穷大
        distances[sorted_indices[0]] = float('inf')
        distances[sorted_indices[-1]] = float('inf')

        # 计算目标值范围（需要解密）
        min_val = combine_shares(
            partial_decrypt(obj_list[sorted_indices[0]][1 + m], share0, pubkey, n_sq),
            partial_decrypt(obj_list[sorted_indices[0]][1 + m], share1, pubkey, n_sq),
            mu, n
        )
        max_val = combine_shares(
            partial_decrypt(obj_list[sorted_indices[-1]][1 + m], share0, pubkey, n_sq),
            partial_decrypt(obj_list[sorted_indices[-1]][1 + m], share1, pubkey, n_sq),
            mu, n
        )

        if max_val - min_val == 0:  # 避免除零错误
            continue

        # 计算中间点的拥挤距离
        for i in range(1, len(sorted_indices) - 1):
            # 解密相邻点的目标值差
            current = combine_shares(
                partial_decrypt(obj_list[sorted_indices[i + 1]][1 + m], share0, pubkey, n_sq),
                partial_decrypt(obj_list[sorted_indices[i + 1]][1 + m], share1, pubkey, n_sq),
                mu, n
            )
            prev = combine_shares(
                partial_decrypt(obj_list[sorted_indices[i - 1]][1 + m], share0, pubkey, n_sq),
                partial_decrypt(obj_list[sorted_indices[i - 1]][1 + m], share1, pubkey, n_sq),
                mu, n
            )
            distances[sorted_indices[i]] += (current - prev) / (max_val - min_val)

    return distances

def binary_tournament_selection(population, obj_list, fronts, crowding_dict, pubkey, share0, share1, mu, n, n_sq):
    """
    NSGA-II的二元锦标赛选择（基于非支配层级和拥挤距离）
    """
    idx1, idx2 = random.sample(range(len(population)), 2)

    # 比较非支配层级
    front1 = next(i for i, front in enumerate(fronts) if idx1 in front)
    front2 = next(i for i, front in enumerate(fronts) if idx2 in front)

    if front1 < front2:
        return idx1
    elif front1 > front2:
        return idx2
    else:
        # 层级相同，比较拥挤距离
        if crowding_dict[idx1] > crowding_dict[idx2]:
            return idx1
        else:
            return idx2


def crossover_and_mutation(parent1, parent2, reach, min_assign):
    """
    针对任务分配矩阵的交叉和变异操作
    """
    T = len(parent1)  # 任务数
    N = len(parent1[0])  # 工人数

    # 单点交叉
    crossover_point = random.randint(1, T - 1)
    child = parent1[:crossover_point] + parent2[crossover_point:]

    # 变异：随机选择一个任务，重新分配工人
    if random.random() < 0.2:  # 变异率20%
        t = random.randint(0, T - 1)

        # 清空当前任务的分配
        child[t] = [0] * N

        # 重新分配：至少分配min_assign个可达工人
        candidates = [i for i in range(N) if reach[t][i] == 1]
        if len(candidates) >= min_assign:
            assigned = random.sample(candidates, min_assign)
            for i in assigned:
                child[t][i] = 1

    return child


def run_nsga2(pubkey, enc_costs, enc_quals, enc_weights,
              share0, share1, mu, n, n_sq, reach,
              num_iter=30, pop_size=20,
              num_tasks=3, min_assign=1,
              fast_mode=False, init_population=None,max_budgets=None):
    """
    NSGA-II主函数（加密环境下）
    """
    N = len(enc_costs[0])  # 工人数
    T = num_tasks

    print("✅ 正在初始化初始种群...")

    if init_population is not None:
        population = init_population
        print("✅ 使用外部传入的初始种群")
    else:
        population = [generate_random_matrix_with_reach(T, N, min_assign, reach)
                      for _ in range(pop_size)]
        print("✅ 默认生成初始种群")

    for gen in tqdm(range(num_iter), desc="🌱 进化轮"):
        print(f"\n🌿 第 {gen + 1}/{num_iter} 轮开始")

        # 评估种群目标
        print("⚡ 开始评估当前种群个体目标值...")
        obj_list = parallel_evaluate_population(
            population=population,
            enc_costs=enc_costs,
            enc_quals=enc_quals,
            enc_weights=enc_weights,
            pubkey=pubkey,
            share0=share0,
            share1=share1,
            mu=mu,
            n=n,
            n_sq=n_sq,
            min_assign=min_assign,
            fast_mode=fast_mode,
            max_budgets=max_budgets  # 传递最大预算
        )
        print("✅ 个体评估完成")

        print("✅ 正在进行非支配排序...")
        fronts = nsga2_non_dominated_sort_enc(obj_list, pubkey, share0, share1, mu, n, n_sq)
        print("✅ 非支配排序完成，前层个体数量:", len(fronts[0]))

        print("📏 计算拥挤距离...")
        # 为每个前沿计算拥挤距离
        crowding_dict = {}
        for front in fronts:
            front_crowding = nsga2_crowding_distance_assignment(obj_list, front, pubkey, share0, share1, mu, n, n_sq)
            crowding_dict.update(front_crowding)
        print("✅ 拥挤距离计算完成")

        # 创建子代
        print("🧬 正在生成下一代个体...")
        offspring = []
        while len(offspring) < pop_size:
            # 选择父代
            parent1_idx = binary_tournament_selection(population, obj_list, fronts, crowding_dict, pubkey, share0,
                                                      share1, mu, n, n_sq)
            parent2_idx = binary_tournament_selection(population, obj_list, fronts, crowding_dict, pubkey, share0,
                                                      share1, mu, n, n_sq)

            # 交叉和变异
            child = crossover_and_mutation(
                population[parent1_idx],
                population[parent2_idx],
                reach,
                min_assign
            )

            offspring.append(child)

        # 合并父代和子代，选择下一代
        combined_pop = population + offspring
        combined_obj = obj_list + parallel_evaluate_population(
            population=offspring,
            enc_costs=enc_costs,
            enc_quals=enc_quals,
            enc_weights=enc_weights,
            pubkey=pubkey,
            share0=share0,
            share1=share1,
            mu=mu,
            n=n,
            n_sq=n_sq,
            min_assign=min_assign,
            fast_mode=fast_mode
        )

        # 基于非支配排序和拥挤距离选择新种群
        new_population = []
        fronts = nsga2_non_dominated_sort_enc(combined_obj, pubkey, share0, share1, mu, n, n_sq)

        for front in fronts:
            if len(new_population) + len(front) <= pop_size:
                # 整个前沿都被选中
                new_population.extend([combined_pop[i] for i in front])
            else:
                # 前沿无法全部选中，按拥挤距离排序选择
                front_crowding = nsga2_crowding_distance_assignment(combined_obj, front, pubkey, share0, share1, mu, n,
                                                                    n_sq)
                sorted_indices = sorted(front, key=lambda i: front_crowding[i], reverse=True)
                remaining = pop_size - len(new_population)
                new_population.extend([combined_pop[i] for i in sorted_indices[:remaining]])
                break

        population = new_population
        print(f"✅ 第 {gen + 1} 轮迭代完成，已生成新种群 ✅")

    # 选择最优解（前沿0中拥挤距离最大的解）
    fronts = nsga2_non_dominated_sort_enc(obj_list, pubkey, share0, share1, mu, n, n_sq)
    front0 = fronts[0]
    crowding_dict = nsga2_crowding_distance_assignment(obj_list, front0, pubkey, share0, share1, mu, n, n_sq)

    best_idx = max(front0, key=lambda i: crowding_dict[i])
    best_x, best_cost, best_qual = obj_list[best_idx]

    print("\n🏆 优化完成，最优任务分配方案如下：")
    for t in range(T):
        print(f"  任务{t + 1}: {best_x[t]}")

    return best_x, best_cost, best_qual, obj_list

# ----------------------------------------------------
# 📌 Crowding Distance
# 输入：obj_values：二维目标值列表，front为个体索引列表
# 输出：返回对应个体的 WCD 距离（np.array，未在front中的为0）
# 参数 alpha 控制权重（目标空间 vs 决策空间）
# ----------------------------------------------------
def compute_weighted_crowding_distance(obj_values, dec_values, front, alpha=0.5):
    n_obj = len(obj_values[0])
    distances = np.zeros(len(obj_values))

    front_obj = [obj_values[i] for i in front]
    front_dec = [dec_values[i] for i in front]

    obj_array = np.array(front_obj)
    dec_array = np.array(front_dec)

    norm_obj = (obj_array - obj_array.min(axis=0)) / (np.ptp(obj_array, axis=0) + 1e-9)
    norm_dec = (dec_array - dec_array.min(axis=0)) / (np.ptp(dec_array, axis=0) + 1e-9)

    dist_obj = np.zeros(len(front))
    dist_dec = np.zeros(len(front))

    for m in range(n_obj):
        idx = np.argsort(norm_obj[:, m])
        dist_obj[idx[0]] = dist_obj[idx[-1]] = float('inf')
        for i in range(1, len(front) - 1):
            dist_obj[idx[i]] += norm_obj[idx[i + 1], m] - norm_obj[idx[i - 1], m]

    for m in range(dec_array.shape[1]):
        idx = np.argsort(norm_dec[:, m])
        dist_dec[idx[0]] = dist_dec[idx[-1]] = float('inf')
        for i in range(1, len(front) - 1):
            dist_dec[idx[i]] += norm_dec[idx[i + 1], m] - norm_dec[idx[i - 1], m]

    for i, idx in enumerate(front):
        distances[idx] = alpha * dist_obj[i] + (1 - alpha) * dist_dec[i]

    return distances


# ---------------------------
# 阈值密钥生成（share0 + share1 = λ）
# ---------------------------
# ===== FastPai 适配层（替代 phe）========================================
import math, random, secrets
from dataclasses import dataclass
try:
    from Crypto.Util.number import getPrime, isPrime
except Exception:
    raise RuntimeError("需要安装 pycryptodome：pip install pycryptodome")

def _gen_PQ_like(p_bits, pprime_bits):
    """
    生成 P = 2*p*p' + 1 为素数；返回 (P, p, p')
    """
    while True:
        p  = getPrime(p_bits)
        p_ = getPrime(pprime_bits) | 1  # 确保奇数
        P  = 2*p*p_ + 1
        if isPrime(P):
            return P, p, p_

@dataclass
class FastPaiPublicKey:
    n: int
    h: int
    scale: int = 10**6  # 若加密 float，可用定点；你当前不需要解密 crowding，可忽略

    def __post_init__(self):
        self.n_sq   = self.n * self.n
        # 给 SMUL/缓存使用的明文上界，用个保守值即可
        self.max_int = self.n // 3

    def encrypt(self, m):
        # 与 phe 接口保持一致：支持 int（必要时也允许 float→定点）
        if isinstance(m, float):
            m = int(round(m * self.scale))
        else:
            m = int(m)
        c1 = pow(1 + self.n, m, self.n_sq)
        r  = secrets.randbits(256)
        c2 = pow(pow(self.h, r, self.n), self.n, self.n_sq)  # (h^r)^N mod N^2
        return FastPaiEncrypted((c1 * c2) % self.n_sq, self)

class FastPaiEncrypted:
    def __init__(self, c, public_key: FastPaiPublicKey):
        self.c = int(c)
        self.public_key = public_key

    # 兼容你代码里的 ciphertext(False)
    def ciphertext(self, raw=False):
        return self.c

    # 密文加：Enc(a) * Enc(b) = Enc(a+b)
    def __add__(self, other):
        if isinstance(other, FastPaiEncrypted):
            return FastPaiEncrypted((self.c * other.c) % self.public_key.n_sq, self.public_key)
        elif isinstance(other, int):
            c_plain = pow(1 + self.public_key.n, other, self.public_key.n_sq)
            return FastPaiEncrypted((self.c * c_plain) % self.public_key.n_sq, self.public_key)
        else:
            raise TypeError(f"Unsupported add with {type(other)}")
    __radd__ = __add__

    # 密文减：Enc(a) * Enc(b)^{-1} = Enc(a-b)
    def __sub__(self, other):
        if isinstance(other, FastPaiEncrypted):
            inv = pow(other.c, -1, self.public_key.n_sq)
            return FastPaiEncrypted((self.c * inv) % self.public_key.n_sq, self.public_key)
        elif isinstance(other, int):
            return self + (-int(other))
        else:
            raise TypeError(f"Unsupported sub with {type(other)}")

    # 标量乘（plaintext k）：Enc(a)^k = Enc(k*a)；k 只能是整数
    def __mul__(self, k):
        if not isinstance(k, int):
            raise TypeError("Only integer scalar supported for ciphertext * k")
        if k == 0:
            # Enc(0) 的一种合法表示（随机性固定为1）
            return FastPaiEncrypted(1, self.public_key)
        if k > 0:
            return FastPaiEncrypted(pow(self.c, k, self.public_key.n_sq), self.public_key)
        # k < 0 ：取逆即可
        pos = pow(self.c, -k, self.public_key.n_sq)
        inv = pow(pos, -1, self.public_key.n_sq)
        return FastPaiEncrypted(inv, self.public_key)
    __rmul__ = __mul__

def fastpai_generate_threshold_keypair(kappa=112):
    """
    生成 FastPai 公钥/阈值私钥分片
    - P=2pp'+1, Q=2qq'+1 为素数
    - N=PQ, alpha=pq, beta=p'q'
    - h = -y^{2beta} (mod N)
    - 私钥指数=2*alpha，并做两份加法分片：share0+share1=2*alpha
    - 为了少改你现有 combine_shares：把 inv_2alpha 放到返回字典的 'mu' 字段
    """
    # 典型设置：让 N≈2048 位。可按需调整位长分配。
    p_bits, pprime_bits = 256, 512
    P, p, p_ = _gen_PQ_like(p_bits, pprime_bits)
    Q, q, q_ = _gen_PQ_like(p_bits, pprime_bits)

    N = P * Q
    alpha = p * q
    beta  = p_ * q_

    # 选 y∈Z*_N
    while True:
        y = random.randrange(2, N - 1)
        if math.gcd(y, N) == 1:
            break
    h = (-pow(y, 2 * beta, N)) % N

    two_alpha = 2 * alpha
    inv_2alpha = pow(two_alpha, -1, N)

    share0 = random.randrange(1, two_alpha)
    share1 = two_alpha - share0   # 保证 share0+share1=2α

    pubkey = FastPaiPublicKey(N, h)

    return {
        'pubkey': pubkey,
        'privkey': None,      # FastPai 不需要暴露 p,q
        'share0': share0,
        'share1': share1,
        'lambda': None,       # 兼容旧字段，占位
        'mu': inv_2alpha,     # ⚠️ 注意：这里存的是 inv_2alpha
        'n': N,
        'n_sq': N * N
    }
# ======================================================================

# 原：def generate_threshold_keypair(): ...（整块删掉）
# 新：直接把 FastPai 的生成器映射为旧名字，调用方不需要改
generate_threshold_keypair = fastpai_generate_threshold_keypair


def lcm(a, b):
    return abs(a * b) // math.gcd(a, b)


# ---------------------------
# 加密函数
# ---------------------------
def encrypt(pubkey, m):
    return pubkey.encrypt(m)


# ---------------------------
# 部分解密函数
# ---------------------------
def partial_decrypt(ciphertext_obj, share_i, pubkey, n_sq):
    c = ciphertext_obj.ciphertext(False)  # 获取原始密文整数
    u = pow(c, share_i, n_sq)
  #  print(f"Partial decrypt result u: {u}")  # 添加调试信息
    return u  # 返回部分解密结果u


# ---------------------------
# 聚合两份部分解密 → 明文
# ---------------------------
def combine_shares(u0, u1, mu, n, max_bits=4096):
    """
    # ⚠️ 安全检查：防止部分解密数值过大
    if u0.bit_length() > max_bits or u1.bit_length() > max_bits:
        print(f"u0 位数: {u0.bit_length()} bit")
        print(f"u1 位数: {u1.bit_length()} bit")
    """
    u = (u0 * u1) % (n * n)  # 合并部分解密结果
    L = (u - 1) // n  # 计算L函数
    m = (L * mu) % n  # 恢复明文
  #  print(f"Combined shares result m: {m}")  # 添加调试信息
    return int(m)




# -------------------------------
# 随机解生成与变异
# -------------------------------
def generate_random_matrix_with_reach(num_tasks, num_workers, min_assign, reach):
    """
    考虑可达性约束的初始化函数：
    - 每个任务分配不少于 min_assign 个工人；
    - 每个工人最多参与 1 个任务；
    - 仅在 reach[t][i] == 1 的条件下才允许分配。
    """
    x = [[0 for _ in range(num_workers)] for _ in range(num_tasks)]
    assigned_workers = set()
    task_assign_count = [0] * num_tasks

    # ✅ 第一步：每个任务分配至少 min_assign 个可达工人
    for t in range(num_tasks):
        candidates = [i for i in range(num_workers) if reach[t][i] == 1 and i not in assigned_workers]
        random.shuffle(candidates)
        for i in candidates[:min_assign]:
            x[t][i] = 1
            assigned_workers.add(i)
            task_assign_count[t] += 1

    # ✅ 第二步：剩余可分配工人（不能分配多个任务）
    for i in range(num_workers):
        if i in assigned_workers:
            continue
        available_tasks = [t for t in range(num_tasks) if reach[t][i] == 1]
        if available_tasks:
            t = random.choice(available_tasks)
            x[t][i] = 1
            assigned_workers.add(i)
            task_assign_count[t] += 1

    return x




# ----------------------------------------------------
# 📌 Fast Non-Dominated Sorting
# 输入：二维列表 obj_values，每个元素是一个个体的 [目标1, 目标2]
# 输出：Pareto 层级列表 fronts，每一层包含个体索引
# ----------------------------------------------------
def is_duplicate(new_sol, pareto_set, pubkey, share0, share1, mu, n, n_sq, epsilon=1e-6):
    """检查 new_sol 是否在 pareto_set 中已存在（去重），需解密目标值后比较。"""
    # new_sol 结构：(x, enc_cost, enc_qual)，提取加密的目标值
    enc_cost_new = new_sol[1]  # 加密的成本
    enc_qual_new = new_sol[2]  # 加密的质量

    # 解密新解的目标值
    dec_cost_new = combine_shares(
        partial_decrypt(enc_cost_new, share0, pubkey, n_sq),
        partial_decrypt(enc_cost_new, share1, pubkey, n_sq),
        mu, n
    )
    dec_qual_new = combine_shares(
        partial_decrypt(enc_qual_new, share0, pubkey, n_sq),
        partial_decrypt(enc_qual_new, share1, pubkey, n_sq),
        mu, n
    )
    new_sol_array = np.array([dec_cost_new, dec_qual_new], dtype=float)  # 确保是浮动类型

    # 遍历已有帕累托解，解密后比较
    for sol in pareto_set:
        # sol 结构：(x, enc_cost, enc_qual)
        enc_cost_sol = sol[1]
        enc_qual_sol = sol[2]

        # 解密已有解的目标值
        dec_cost_sol = combine_shares(
            partial_decrypt(enc_cost_sol, share0, pubkey, n_sq),
            partial_decrypt(enc_cost_sol, share1, pubkey, n_sq),
            mu, n
        )
        dec_qual_sol = combine_shares(
            partial_decrypt(enc_qual_sol, share0, pubkey, n_sq),
            partial_decrypt(enc_qual_sol, share1, pubkey, n_sq),
            mu, n
        )
        sol_array = np.array([dec_cost_sol, dec_qual_sol], dtype=float)  # 确保是浮动类型

        # 计算欧氏距离判断是否重复
        distance = np.linalg.norm(new_sol_array - sol_array)  # 这里确保是浮动类型
        if distance < epsilon:
            return True
    return False

def fast_non_dominated_sort_enc(obj_list, pubkey, share0, share1, mu, n, n_sq):
    """
    使用密文比较完成非支配排序过程，适用于 encrypted obj_list。
    每个 obj_list[i] 是加密后的目标值列表。
    返回：fronts（List[List[int]]）多层非支配解索引
    """
    S = {}         # 支配集合
    n_dom = {}     # 被支配次数
    fronts = [[]]  # 非支配层

    for p in range(len(obj_list)):
        S[p] = []
        n_dom[p] = 0
        for q in range(len(obj_list)):
            if p == q:
                continue
            if dominates_enc(obj_list[p], obj_list[q], pubkey, share0, share1, mu, n, n_sq):
                S[p].append(q)
            elif dominates_enc(obj_list[q], obj_list[p], pubkey, share0, share1, mu, n, n_sq):
                n_dom[p] += 1
        if n_dom[p] == 0:
            fronts[0].append(p)

    i = 0
    while len(fronts[i]) > 0:
        next_front = []
        for p in fronts[i]:
            for q in S[p]:
                n_dom[q] -= 1
                if n_dom[q] == 0:
                    next_front.append(q)
        i += 1
        fronts.append(next_front)

    # 去掉最后一个空层
    if len(fronts[-1]) == 0:
        fronts.pop()

    # 去重：检查每个解是否已存在（补充解密参数）
    unique_fronts = []
    for front in fronts:
        unique_front = []
        for idx in front:
            # 传入解密所需的参数给 is_duplicate
            if not is_duplicate(
                obj_list[idx],
                [obj_list[i] for i in unique_front],  # 已选的解列表
                pubkey, share0, share1, mu, n, n_sq  # 新增：解密参数
            ):
                unique_front.append(idx)
        unique_fronts.append(unique_front)

    return unique_fronts


def compute_weighted_crowding_distance(obj_list, population, front_indices,
                                       pubkey, share0, share1, mu, n, n_sq,
                                       alpha=0.5):
    """
    使用加权替代策略计算密文拥塞距离（无需排序）：
    EncCrowding = α × (enc_max_cost - enc_cost) + (1 - α) × enc_qual
    """
    enc_costs = [obj_list[i][1] for i in front_indices]
    enc_quals = [obj_list[i][2] for i in front_indices]

    # 👑 先选出 enc_max_cost
    max_cost_idx = select_argmax_enc(list(zip(front_indices, enc_costs)),
                                     pubkey, share0, share1, mu, n, n_sq)
    enc_max_cost = obj_list[max_cost_idx][1]

    crowding_dict = {}

    for i in front_indices:
        enc_cost = obj_list[i][1]
        enc_qual = obj_list[i][2]

        # ➗ crowding = α*(max_cost - cost) + (1-α)*qual
        enc_diff = enc_max_cost - enc_cost
        enc_part1 = enc_diff * alpha
        enc_part2 = enc_qual * (1 - alpha)
        enc_crowding = enc_part1 + enc_part2

        crowding_dict[i] = enc_crowding

    return crowding_dict


#综合使用密文选择 + 随机补全：
def integrate_elite_selection(population, crowding_dict, k, pubkey, share0, share1, mu, n, n_sq):
    """
    综合使用密文排序与随机补全进行精英个体选择。
    返回：elite_indices 精英个体索引列表
    """
    elite_indices = select_topk_by_enc_value(crowding_dict, k, pubkey, share0, share1, mu, n, n_sq)
    if len(elite_indices) < k:
        supplement = select_random_if_empty(crowding_dict, k - len(elite_indices))
        elite_indices += supplement
    return elite_indices


def select_best_by_custom_score(obj_list, share0, share1, pubkey, mu, n, n_sq, alpha=0.7, beta=0.3, eps=1e-6):
    """
    基于自定义评分函数 Score = (alpha * quality) / (beta * cost) 选择最优个体
    """
    best_idx = None
    best_score = float('-inf')

    for i, (_, enc_cost, enc_qual) in enumerate(obj_list):
        dec_cost = combine_shares(
            partial_decrypt(enc_cost, share0, pubkey, n_sq),
            partial_decrypt(enc_cost, share1, pubkey, n_sq),
            mu, n
        )
        dec_qual = combine_shares(
            partial_decrypt(enc_qual, share0, pubkey, n_sq),
            partial_decrypt(enc_qual, share1, pubkey, n_sq),
            mu, n
        )

        score = (dec_qual * alpha) / (dec_cost * beta + eps)
        if score > best_score:
            best_score = score
            best_idx = i

    return best_idx


# -------------------------------
# 密文目标函数评估
# -------------------------------
def evaluate_cost_stable_smulg(x, enc_costs, enc_weights, pubkey, share0, share1, mu, n, n_sq):
    """
    现在的成本 = ∑ Enc(cost_{t,i})，不再与 enc_weights 相乘
    （保留 enc_weights 形参是为了不改调用处签名，但这里不用它）
    """
    T = len(x)
    N = len(x[0])

    total_cost = None
    for t in range(T):
        for i in range(N):
            if x[t][i] == 1:
                enc_c = enc_costs[t][i]             # Enc(cost_{t,i})
                total_cost = enc_c if total_cost is None else (total_cost + enc_c)
    return total_cost if total_cost else pubkey.encrypt(0)




def evaluate_quality_stable(x, enc_quals, enc_weights,
                            pubkey, share0, share1, mu, n, n_sq):
    """
    现在的质量 = ∑ Enc(quality_{t,i} * weight_i)
    使用 secure_multiply 做密文乘法，再同态加法相加
    """
    T = len(x)
    N = len(x[0])

    total_qual = None
    for t in range(T):
        for i in range(N):
            if x[t][i] == 1:
                enc_q = enc_quals[t][i]              # Enc(quality_{t,i})
                enc_w = enc_weights[i]               # Enc(weight_i)
                enc_qw = secure_multiply(enc_q, enc_w, pubkey, share0, share1, mu, n, n_sq)
                total_qual = enc_qw if total_qual is None else (total_qual + enc_qw)
    return total_qual if total_qual else pubkey.encrypt(0)

def evaluate_individual_parallel(x, enc_costs, enc_quals, enc_weights,
                                 pubkey, share0, share1, mu, n, n_sq,
                                 min_assign, fast_mode=False, max_budgets=None):
    """
    新增 max_budgets: List[int] 或 None
      - 若提供，则对每个任务 t，累计该任务下被分配工人的“明文成本”并与 max_budgets[t] 比较
      - 超预算则该解判为无效（高成本、零质量）
    其他约束：
      - 每个任务至少 min_assign
      - 每个工人最多参与 1 个任务
    """
    T = len(x)
    N = len(x[0])
    valid = True

    # 任务最小分配
    for t in range(T):
        if sum(x[t]) < min_assign:
            valid = False
            break

    # 工人唯一性
    if valid:
        for i in range(N):
            if sum(x[t][i] for t in range(T)) > 1:
                valid = False
                break

    # 任务预算（如提供）
    if valid and max_budgets is not None:
        for t in range(T):
            dec_task_cost = 0
            for i in range(N):
                if x[t][i] == 1:
                    dec_task_cost += combine_shares(
                        partial_decrypt(enc_costs[t][i], share0, pubkey, n_sq),
                        partial_decrypt(enc_costs[t][i], share1, pubkey, n_sq),
                        mu, n
                    )
            if dec_task_cost > max_budgets[t]:
                valid = False
                break

    if fast_mode:
        cost = sum(i * x[t][i] for t in range(T) for i in range(N))
        qual = sum(10 * x[t][i] for t in range(T) for i in range(N))
        return {"x": x, "enc_cost": pubkey.encrypt(cost), "enc_qual": pubkey.encrypt(qual), "key": str(x)}

    if not valid:
        enc_cost = pubkey.encrypt(999999)
        enc_qual = pubkey.encrypt(0)
    else:
        enc_cost = evaluate_cost_stable_smulg(x, enc_costs, enc_weights, pubkey, share0, share1, mu, n, n_sq)
        enc_qual = evaluate_quality_stable(x, enc_quals, enc_weights, pubkey, share0, share1, mu, n, n_sq)

    return {"x": x, "enc_cost": enc_cost, "enc_qual": enc_qual, "key": str(x)}


def parallel_evaluate_population(population, enc_costs, enc_quals, enc_weights,
                                 pubkey, share0, share1, mu, n, n_sq,
                                 min_assign, fast_mode=False,
                                 use_cache=True, max_budgets=None):
    from concurrent.futures import ProcessPoolExecutor, as_completed
    from tqdm import tqdm
    import multiprocessing

    # ✅ 初始化缓存
    global decrypt_cache
    if 'decrypt_cache' not in globals():
        decrypt_cache = {}

    results = []
    with ProcessPoolExecutor(max_workers=multiprocessing.cpu_count()) as executor:
        futures = []
        for x in population:
            key = str(x)
            if use_cache and key in decrypt_cache:
                results.append(decrypt_cache[key])
            else:
                futures.append(executor.submit(
                    evaluate_individual_parallel, x,
                    enc_costs, enc_quals, enc_weights,
                    pubkey, share0, share1, mu, n, n_sq,
                    min_assign, fast_mode, max_budgets
                ))

        for f in tqdm(as_completed(futures), total=len(futures), desc="⚡并行评估中"):
            result = f.result()
            results.append((result["x"], result["enc_cost"], result["enc_qual"]))
            if use_cache:
                decrypt_cache[result["key"]] = (result["x"], result["enc_cost"], result["enc_qual"])

    return results




def dominates_enc(ind1, ind2, pubkey, share0, share1, mu, n, n_sq):
    enc_cost1, enc_qual1 = ind1[1], ind1[2]
    enc_cost2, enc_qual2 = ind2[1], ind2[2]

    comp1 = secure_compare(enc_cost1, enc_cost2, pubkey, share0, share1, mu, n, n_sq)
    comp2 = secure_compare(enc_qual2, enc_qual1, pubkey, share0, share1, mu, n, n_sq)

    u01 = partial_decrypt(comp1, share0, pubkey, n_sq)
    u02 = partial_decrypt(comp2, share0, pubkey, n_sq)
    u11 = partial_decrypt(comp1, share1, pubkey, n_sq)
    u12 = partial_decrypt(comp2, share1, pubkey, n_sq)

    b1 = combine_shares(u01, u11, mu, n)  # cost1 < cost2
    b2 = combine_shares(u02, u12, mu, n)  # qual1 > qual2

    return b1 == 1 and b2 == 1


#使用密文比较获取最大前k个
def select_topk_by_enc_value(enc_dict, k, pubkey, share0, share1, mu, n, n_sq):
    """
    从加密值字典中选择 Top-k 的索引（基于密文比较）。
    enc_dict: {index: Enc(value)}
    返回：长度为 ≤ k 的索引列表（按密文最大值优先）
    """
    selected = []
    remaining = list(enc_dict.keys())

    while len(selected) < k and remaining:
        max_idx = remaining[0]
        for i in remaining[1:]:
            comp = secure_compare(enc_dict[i], enc_dict[max_idx], pubkey, share0, share1, mu, n, n_sq)
            u0 = partial_decrypt(comp, share0, pubkey, n_sq)
            u1 = partial_decrypt(comp, share1, pubkey, n_sq)
            result = combine_shares(u0, u1, mu, n)
            if result == 1:
                max_idx = i

        selected.append(max_idx)
        remaining.remove(max_idx)

    return selected


#密文Argmax选择器
def select_argmax_enc(index_value_pairs, pubkey, share0, share1, mu, n, n_sq):
    """
    从 (index, enc_value) 列表中选出最大值对应索引
    """
    max_index, max_enc = index_value_pairs[0]

    for idx, enc in index_value_pairs[1:]:
        enc_cmp = secure_compare(enc, max_enc, pubkey, share0, share1, mu, n, n_sq)
        u0 = partial_decrypt(enc_cmp, share0, pubkey, n_sq)
        u1 = partial_decrypt(enc_cmp, share1, pubkey, n_sq)
        bit = combine_shares(u0, u1, mu, n)
        if bit == 1:
            max_index, max_enc = idx, enc

    return max_index


#字典为空时随机返回 k 个键（用于补全）
def select_random_if_empty(enc_dict, k):
    """如果密文字典为空，则随机选择 k 个索引作为后备方案。"""
    all_indices = list(enc_dict.keys())
    if not all_indices:
        print("⚠️ 后备选择：字典为空，返回空列表")
        return []
    return random.sample(all_indices, min(k, len(all_indices)))


def simulate_worker_upload(num_tasks=10, num_workers=25):
    # 校验参数是否与固定数据匹配（5个任务，15个工人）
    if num_tasks != 10 or num_workers != 25:
        raise ValueError("固定数据仅支持num_tasks=5和num_workers=15")

    keys = generate_threshold_keypair()
    pubkey = keys['pubkey']
    share0, share1 = keys['share0'], keys['share1']
    mu, n, n_sq = keys['mu'], keys['n'], keys['n_sq']

    # --------------------------
    # 固定的工人成本权重向量（25 名工人）
    # --------------------------
    raw_weights = [2, 2, 5, 1, 1, 4, 3, 5, 3, 2, 3, 3, 4, 2, 3, 5, 2, 4, 1, 3, 4, 2, 5, 3, 2]

    # --------------------------
    # 固定的工人成本矩阵 (25 个工人 × 10 个任务)
    # 注意：这里按“工人×任务”填写；下一行会转置成“任务×工人”使用
    # --------------------------
    raw_costs = [
        [26, 14, 21, 29, 30, 18, 22, 24, 19, 21],  # 工人1
        [15, 29, 11, 11, 19, 17, 24, 20, 18, 16],  # 工人2
        [12, 28, 23, 17, 22, 16, 19, 21, 15, 20],  # 工人3
        [30, 15, 10, 18, 10, 20, 13, 14, 22, 17],  # 工人4
        [14, 21, 15, 19, 26, 12, 18, 16, 20, 23],  # 工人5
        [21, 30, 29, 20, 15, 14, 20, 22, 25, 19],  # 工人6
        [11, 13, 11, 10, 11, 12, 16, 18, 14, 13],  # 工人7
        [22, 10, 10, 28, 23, 19, 21, 24, 20, 18],  # 工人8
        [30, 23, 20, 10, 19, 18, 24, 22, 21, 25],  # 工人9
        [27, 13, 18, 10, 16, 15, 20, 19, 17, 14],  # 工人10
        [20, 13, 11, 20, 26, 14, 17, 21, 16, 18],  # 工人11
        [12, 23, 28, 13, 24, 16, 20, 18, 22, 21],  # 工人12
        [23, 13, 12, 23, 29, 21, 18, 20, 24, 22],  # 工人13
        [27, 24, 23, 28, 30, 22, 25, 26, 27, 23],  # 工人14
        [26, 15, 12, 29, 12, 18, 21, 17, 16, 19],  # 工人15
        [18, 16, 20, 22, 14, 17, 19, 20, 18, 21],  # 工人16
        [19, 21, 17, 15, 18, 13, 20, 22, 23, 16],  # 工人17
        [16, 19, 22, 14, 20, 18, 23, 21, 19, 24],  # 工人18
        [24, 17, 15, 21, 16, 19, 18, 23, 20, 22],  # 工人19
        [17, 20, 19, 16, 22, 20, 21, 18, 17, 23],  # 工人20
        [25, 18, 16, 22, 19, 17, 20, 21, 24, 18],  # 工人21
        [14, 22, 20, 19, 17, 16, 18, 20, 22, 21],  # 工人22
        [21, 19, 24, 17, 23, 20, 22, 25, 19, 20],  # 工人23
        [20, 16, 18, 23, 21, 19, 17, 22, 18, 24],  # 工人24
        [22, 21, 17, 20, 18, 22, 19, 23, 21, 17],  # 工人25
    ]
    # 转置成本矩阵（原数据是工人×任务，需要转为任务×工人）
    raw_costs = [[raw_costs[i][t] for i in range(num_workers)] for t in range(num_tasks)]

    # --------------------------
    # 固定的任务质量矩阵 (10 个任务 × 25 个工人)
    # 每行对应一个任务，对应 25 名工人的质量评分（范围 5~15）
    # --------------------------
    raw_quals = [
        [13, 8, 12, 7, 8, 9, 7, 15, 9, 10, 6, 15, 8, 6, 8, 11, 12, 9, 7, 10, 12, 9, 14, 11, 10],  # 任务1
        [5, 13, 13, 15, 6, 9, 6, 13, 11, 15, 15, 8, 14, 7, 6, 10, 9, 12, 8, 11, 13, 10, 12, 9, 14],  # 任务2
        [6, 11, 6, 12, 8, 14, 10, 15, 10, 7, 10, 12, 11, 11, 13, 9, 12, 10, 7, 14, 9, 12, 11, 10, 13],  # 任务3
        [6, 9, 13, 13, 10, 11, 6, 5, 6, 13, 8, 8, 14, 5, 8, 12, 11, 9, 10, 7, 12, 7, 11, 9, 10],  # 任务4
        [10, 15, 13, 14, 12, 11, 6, 11, 15, 5, 15, 6, 6, 6, 7, 13, 12, 10, 9, 8, 11, 12, 10, 13, 9],  # 任务5
        [9, 12, 11, 10, 7, 13, 8, 12, 9, 11, 10, 9, 12, 8, 10, 14, 11, 12, 9, 13, 10, 11, 12, 9, 12],  # 任务6
        [8, 10, 12, 9, 11, 12, 9, 13, 10, 9, 11, 12, 10, 12, 9, 11, 10, 13, 8, 12, 9, 10, 11, 12, 10],  # 任务7
        [12, 9, 11, 10, 9, 13, 7, 12, 11, 10, 9, 12, 11, 9, 10, 13, 12, 11, 8, 12, 10, 12, 9, 11, 13],  # 任务8
        [11, 12, 9, 8, 10, 12, 8, 11, 12, 9, 10, 11, 12, 10, 9, 12, 11, 10, 9, 13, 12, 9, 12, 10, 11],  # 任务9
        [10, 11, 13, 9, 12, 11, 9, 10, 11, 12, 9, 13, 10, 11, 12, 9, 13, 11, 10, 12, 11, 12, 10, 13, 9],  # 任务10
    ]

    # --------------------------
    # 固定的工人位置与最大距离（25 名工人）
    # --------------------------
    raw_worker_locs = [
        (7, 30), (34, 24), (15, 16), (24, 36), (46, 6),
        (42, 26), (8, 6), (0, 9), (17, 8), (37, 32),
        (36, 9), (18, 30), (45, 20), (26, 7), (25, 46),
        (12, 22), (33, 14), (5, 27), (28, 18), (9, 41),
        (14, 34), (31, 5), (8, 44), (40, 15), (19, 27),
    ]
    raw_max_dists = [
        41, 35, 47, 47, 27, 35, 22, 39, 24, 27,
        43, 32, 28, 35, 42, 30, 33, 36, 29, 40,
        34, 31, 45, 38, 33,
    ]

    # --------------------------
    # 固定的任务位置（10 个任务）
    # --------------------------
    raw_task_locs = [
        (21, 49), (43, 13), (3, 12), (5, 3), (23, 39),
        (15, 25), (38, 28), (30, 20), (12, 40), (44, 30),
    ]
    # 加密数据
    enc_weights = [encrypt(pubkey, w) for w in raw_weights]
    enc_costs = [[encrypt(pubkey, raw_costs[t][i]) for i in range(num_workers)] for t in range(num_tasks)]
    enc_quals = [[encrypt(pubkey, raw_quals[t][i]) for i in range(num_workers)] for t in range(num_tasks)]
    enc_task_locs = [(encrypt(pubkey, x), encrypt(pubkey, y)) for x, y in raw_task_locs]
    enc_worker_locs = [(encrypt(pubkey, x), encrypt(pubkey, y)) for x, y in raw_worker_locs]
    enc_max_dists = [encrypt(pubkey, d) for d in raw_max_dists]

    # 保存加密数据
    with open("enc_worker_data.pkl", "wb") as f:
        pickle.dump({
            "pubkey": pubkey,
            "costs": enc_costs,
            "quals": enc_quals,
            "raw_costs": raw_costs,
            "raw_quals": raw_quals,
            "raw_weights": raw_weights,
            "enc_weights": enc_weights,
            "raw_task_locs": raw_task_locs,
            "raw_worker_locs": raw_worker_locs,
            "raw_max_dists": raw_max_dists,
            "enc_task_locs": enc_task_locs,
            "enc_worker_locs": enc_worker_locs,
            "enc_max_dists": enc_max_dists
        }, f)

    # 保存密钥分片
    with open("threshold_key_shares.pkl", "wb") as f:
        pickle.dump({
            "share0": share0,
            "share1": share1,
            "mu": mu,
            "n": n,
            "n_sq": n_sq
        }, f)

    print("✅ 多任务加密数据上传成功（使用固定数据）")




# -----------------------------
# 工人数据准备 + 平台密钥加载
# -----------------------------
def prepare_data(num_tasks, num_workers):
    # 模拟上传数据
    simulate_worker_upload(num_tasks=num_tasks, num_workers=num_workers)

    # 加载加密数据与密钥
    with open('enc_worker_data.pkl', 'rb') as f:
        data = pickle.load(f)
    with open('threshold_key_shares.pkl', 'rb') as f:
        key_parts = pickle.load(f)

    # 定义每个任务的最大预算（10 个任务）
    max_budgets = [100, 150, 200, 120, 180, 200, 190, 170, 210, 220]

    return (
        # 🔐 加密部分
        data['pubkey'], data['costs'], data['quals'], data['enc_weights'],
        data['enc_task_locs'], data['enc_worker_locs'], data['enc_max_dists'],

        # 🧩 明文部分
        data['raw_costs'], data['raw_quals'], data['raw_weights'],
        data['raw_task_locs'], data['raw_worker_locs'], data['raw_max_dists'],

        # 🔑 密钥部分
        key_parts['share0'], key_parts['share1'], key_parts['mu'], key_parts['n'], key_parts['n_sq'],

        # 额外返回的预算
        max_budgets  # 返回每个任务的最大预算
    )



# ✅ 增强版：MOEO-WCD 引导浓度机制模块
# 💡 本文件应与你已有的 run_moeo_wcd 函数一起使用，替代其中变异生成子代部分
def generate_ceq_pool(parent, population, obj_list, strategy="mixed",
                      pubkey=None, share0=None, share1=None, mu=None, n=None, n_sq=None,
                      parent_idx=None):
    """
    生成平衡池浓度集合（支持 parent_idx 避免 .index 报错）
    """
    Ceq_pool = []

    if strategy in ["random", "mixed"]:
        neighbors = random.sample(population, min(5, len(population)))
        Ceq_pool.extend(neighbors)
        avg = np.mean([np.array(nei).flatten() for nei in neighbors], axis=0).reshape(np.array(parent).shape)
        Ceq_pool.append(avg.tolist())

    if strategy in ["decision", "mixed"]:
        parent_vec = np.array(parent).flatten()
        distances = [(np.sum(np.abs(np.array(p).flatten() - parent_vec)), p) for p in population]
        distances.sort(reverse=True)
        neighbors = [p for _, p in distances[:5]]
        Ceq_pool.extend(neighbors)
        avg = np.mean([np.array(nei).flatten() for nei in neighbors], axis=0).reshape(np.array(parent).shape)
        Ceq_pool.append(avg.tolist())

    if strategy in ["objective", "mixed"] and pubkey:
        if parent_idx is None:
            raise ValueError("parent_idx must be provided explicitly to avoid .index() errors.")

        parent_obj = obj_list[parent_idx]

        def dec_obj(o):
            c = combine_shares(partial_decrypt(o[1], share0, pubkey, n_sq),
                               partial_decrypt(o[1], share1, pubkey, n_sq), mu, n)
            q = combine_shares(partial_decrypt(o[2], share0, pubkey, n_sq),
                               partial_decrypt(o[2], share1, pubkey, n_sq), mu, n)
            return c, q

        parent_c, parent_q = dec_obj(parent_obj)
        distances = []
        for p, o in zip(population, obj_list):
            c, q = dec_obj(o)
            d = abs(c - parent_c) + abs(q - parent_q)
            distances.append((d, p))
        distances.sort(reverse=True)
        neighbors = [p for _, p in distances[:5]]
        Ceq_pool.extend(neighbors)
        avg = np.mean([np.array(nei).flatten() for nei in neighbors], axis=0).reshape(np.array(parent).shape)
        Ceq_pool.append(avg.tolist())

    return Ceq_pool

def update_with_ceq(parent, ceq_pool, lambda_=1.0):
    """
    简化浓度引导更新：parent + λ * (Ceq - parent) + noise → binary
    """
    selected = random.choice(ceq_pool)
    parent_np = np.array(parent)
    ceq_np = np.array(selected)
    delta = ceq_np - parent_np
    noise = np.random.uniform(-0.05, 0.05, size=parent_np.shape)
    updated = parent_np + lambda_ * delta + noise
    updated = (updated > 0.5).astype(int).tolist()
    return updated



# -------------------------------
# 主优化器（支持阈值解密 + 分配约束）
# -------------------------------
def run_moeo_wcd(pubkey, enc_costs, enc_quals, enc_weights,
                 share0, share1, mu, n, n_sq, reach,
                 num_iter=30, pop_size=20,
                 num_tasks=3, min_assign=2,
                 fast_mode=False,init_population=None):
    """
    多任务-多目标优化主函数，支持并行评估 + 快速评估 + tqdm 进度。
    增加调试日志用于确认程序运行状态。
    """
    N = len(enc_costs[0])  # 工人数
    T = num_tasks

    print("✅ 正在初始化初始种群...")

    if init_population is not None:
        population = init_population  # ✅ 使用外部传入的种群
        print("✅ 使用外部传入的初始种群")
    else:
        population = [generate_random_matrix_with_reach(T, N, min_assign, reach)
                      for _ in range(pop_size)]
        print("✅ 默认生成初始种群")

    print("✅ 初始种群构造完成，共", len(population), "个体")
    # ✅ 打印每个个体
    for idx, individual in enumerate(population):
        print(f"\n🌱 个体 {idx + 1}:")
        for t_idx, task_row in enumerate(individual):
            print(f"  任务 {t_idx + 1}: {task_row}")

    for gen in tqdm(range(num_iter), desc="🌱 进化轮"):
        print(f"\n🌿 第 {gen + 1}/{num_iter} 轮开始")

        # 评估种群目标
        print("⚡ 开始评估当前种群个体目标值...")
        obj_list = parallel_evaluate_population(
            population=population,
            enc_costs=enc_costs,
            enc_quals=enc_quals,
            enc_weights=enc_weights,
            pubkey=pubkey,
            share0=share0,
            share1=share1,
            mu=mu,
            n=n,
            n_sq=n_sq,
            min_assign=min_assign,
            fast_mode=fast_mode
        )
        print("✅ 个体评估完成")

        print("✅ 正在进行非支配排序...")
        fronts = fast_non_dominated_sort_enc(obj_list, pubkey, share0, share1, mu, n, n_sq)
        print("✅ 非支配排序完成，前层个体数量:", len(fronts[0]))

        print("📏 计算加权拥塞距离...")
        crowding = compute_weighted_crowding_distance(obj_list, population, fronts[0],
                                                      pubkey, share0, share1, mu, n, n_sq,
                                                      alpha=0.5)
        print("✅ 拥塞距离计算完成")

        print("🎯 正在选择精英个体...")
        elite_indices = integrate_elite_selection(
            population, crowding, pop_size // 2, pubkey, share0, share1, mu, n, n_sq
        )
        print(f"✅ 精英选择完成，选中 {len(elite_indices)} 个体")

        print("🧬 正在生成下一代个体...")
        unique_new_population = []
        retry_counter = 0
        max_retry = pop_size * 5

        while len(unique_new_population) < pop_size and retry_counter < max_retry:
            parent_idx = random.randrange(len(population))
            parent = population[parent_idx]

            ceq_pool = generate_ceq_pool(parent, population, obj_list, strategy="mixed",
                                         pubkey=pubkey, share0=share0, share1=share1,
                                         mu=mu, n=n, n_sq=n_sq, parent_idx=parent_idx)

            child = update_with_ceq(parent, ceq_pool)

            if child not in unique_new_population:
                unique_new_population.append(child)
                print(f"🧪 子代 {len(unique_new_population)} / {pop_size} 生成完成")
            else:
                retry_counter += 1

        if retry_counter >= max_retry:
            print("⚠️ 警告：子代生成重复太多，已达最大尝试次数")

        print(f"✅ 第 {gen + 1} 轮迭代完成，已生成新种群 ✅")

        # 更新种群
        population = unique_new_population

    print("🎯 正在从最终个体中选择最优解...")
    best_idx = select_best_by_custom_score(
        obj_list=obj_list,
        share0=share0,
        share1=share1,
        pubkey=pubkey,
        mu=mu,
        n=n,
        n_sq=n_sq,
        alpha=0.7,
        beta=0.3
    )

    best_x, best_cost, best_qual = obj_list[best_idx]

    print("\n🏆 优化完成，最优任务分配方案如下：")
    for t in range(T):
        print(f"  任务{t + 1}: {best_x[t]}")

    return best_x, best_cost, best_qual, obj_list



#手动开方
def isqrt(n):
    x = n
    y = (x + 1) // 2
    while y < x:
        x = y
        y = (x + n // x) // 2
    return x

#后加的密文下乘法（SMUL）
def secure_multiply(enc_x, enc_y, pubkey, share0, share1, mu, n, n_sq):
    """
    多进程安全版本，使用本地加载缓存，不依赖 offline_ptr
    """
    cache = get_next_cache_process_safe(pubkey=pubkey)

    r1 = cache['r1']
    r2 = cache['r2']
    enc_r1 = cache['enc_r1']
    enc_r2 = cache['enc_r2']
    enc_r1r2 = cache['enc_r1r2']

    enc_x_r2 = enc_x * r2 * -1
    enc_y_r1 = enc_y * r1 * -1

    enc_x_ = enc_x + enc_r1
    enc_y_ = enc_y + enc_r2

    u0 = partial_decrypt(enc_y_, share0, pubkey, n_sq)
    u1 = partial_decrypt(enc_y_, share1, pubkey, n_sq)
    y_plus_r2 = combine_shares(u0, u1, mu, n)

    enc_xy_noisy = enc_x_ * y_plus_r2
    result = enc_xy_noisy + enc_x_r2 + enc_y_r1 + (enc_r1r2 * -1)
    return result


#后加的密文比较协议SCMP

def secure_compare(enc_x, enc_y, pubkey, share0, share1, mu, n, n_sq):
    """
    多进程安全版本，使用本地加载缓存
    """
    cache = get_next_cache_process_safe(pubkey=pubkey)

    r = cache['r1']
    r_dash = random.randint(n // 5, n // 4)

    enc_diff_base = enc_y - enc_x + pubkey.encrypt(1)
    enc_r_diff = enc_diff_base * r
    enc_full = enc_r_diff + pubkey.encrypt(r_dash)

    u0 = partial_decrypt(enc_full, share0, pubkey, n_sq)
    u1 = partial_decrypt(enc_full, share1, pubkey, n_sq)
    d = combine_shares(u0, u1, mu, n)

    return cache['enc_1'] if d > r_dash else cache['enc_0']

#计算密文下曼哈顿距离（模拟绝对值）
def secure_manhattan_distance(task_loc_enc, worker_loc_enc, pubkey, share0, share1, mu, n, n_sq):
    """
    输入：任务与工人的位置 Enc(x, y)
    输出：Enc(|x_t - x_i| + |y_t - y_i|)
    """
    enc_xt, enc_yt = task_loc_enc
    enc_xi, enc_yi = worker_loc_enc

    # 差值：dx = xt - xi，dy = yt - yi
    enc_dx = enc_xt - enc_xi
    enc_dy = enc_yt - enc_yi

    # 模拟绝对值：abs(a) = max(a, -a)
    def abs_enc(enc_a):
        enc_neg_a = enc_a * -1  # ✅ 替代 -enc_a
        cmp = secure_compare(enc_a, enc_neg_a, pubkey, share0, share1, mu, n, n_sq)
        u0 = partial_decrypt(cmp, share0, pubkey, n_sq)
        u1 = partial_decrypt(cmp, share1, pubkey, n_sq)
        b = combine_shares(u0, u1, mu, n)
        return enc_neg_a if b == 1 else enc_a  # b==1 → 原本负数，返回 -a，否则返回 a

    abs_dx = abs_enc(enc_dx)
    abs_dy = abs_enc(enc_dy)
    return abs_dx + abs_dy  # 曼哈顿距离密文

#判断是否小于等于最大可接受距离
def check_reachability_matrix(enc_task_locs, enc_worker_locs, enc_max_dists,
                               pubkey, share0, share1, mu, n, n_sq):
    """
    返回 reach[t][i] = 1 表示工人 i 可参与任务 t
    """
    num_tasks = len(enc_task_locs)
    num_workers = len(enc_worker_locs)
    reach = [[0 for _ in range(num_workers)] for _ in range(num_tasks)]

    for t in range(num_tasks):
        for i in range(num_workers):
            enc_dist = secure_manhattan_distance(enc_task_locs[t], enc_worker_locs[i],
                                                 pubkey, share0, share1, mu, n, n_sq)
            cmp = secure_compare(enc_dist, enc_max_dists[i], pubkey, share0, share1, mu, n, n_sq)
            u0 = partial_decrypt(cmp, share0, pubkey, n_sq)
            u1 = partial_decrypt(cmp, share1, pubkey, n_sq)
            b = combine_shares(u0, u1, mu, n)
            reach[t][i] = 1 if b == 1 else 0  # dist ≤ max_dist ⇒ 可达

    return reach

def check_task_reachability(reach):
    for t_idx, row in enumerate(reach):
        if sum(row) == 0:
            raise ValueError(f"❌ 错误：任务 {t_idx + 1} 无法被任何工人完成（reach 全为0）")



def generate_single_cache(pubkey):
    """
    生成一组加密扰动元组：r1, r2, Enc(r1), Enc(r2), Enc(r1*r2), Enc(0), Enc(1)
    """
    max_val = pubkey.max_int // 4
    while True:
        r1 = random.randint(1, 10000)
        r2 = random.randint(1, 10000)
        if r1 * r2 <= max_val:
            break

    return {
        'r1': r1,
        'r2': r2,
        'enc_r1': pubkey.encrypt(r1),
        'enc_r2': pubkey.encrypt(r2),
        'enc_r1r2': pubkey.encrypt(r1 * r2),
        'enc_0': pubkey.encrypt(0),
        'enc_1': pubkey.encrypt(1)
    }

def generate_offline_cache(pubkey, num_sets=600, num_workers=16):
    """
    多线程版本：生成 num_sets 个离线扰动缓存，并保存为 offline_cache.pkl
    """
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        cache_list = list(executor.map(lambda _: generate_single_cache(pubkey), range(num_sets)))

    with open("offline_cache.pkl", "wb") as f:
        pickle.dump(cache_list, f)

    print(f"✅ 多线程生成完成，共 {num_sets} 组扰动元组，使用 {num_workers} 线程")



#在线阶段加载缓存
def load_offline_cache():
    """
    加载本地保存的 offline 缓存扰动元组
    :return: list of dict
    """
    with open("offline_cache.pkl", "rb") as f:
        return pickle.load(f)



def get_next_cache_process_safe(pubkey=None, refill_num=100):
    """
    多进程安全版本的 get_next_cache（子进程中使用，不依赖全局变量）
    每次调用从 offline_cache.pkl 中加载一组
    """
    if not os.path.exists("offline_cache.pkl"):
        if pubkey is None:
            raise ValueError("❌ 无 offline_cache.pkl 且未提供 pubkey")
        generate_offline_cache(pubkey, num_sets=refill_num)

    with open("offline_cache.pkl", "rb") as f:
        cache_list = pickle.load(f)

    # 每次随机返回一组（可选优化：循环指针或共享内存）
    return random.choice(cache_list)

import os, json
from datetime import datetime

import json
import time
from datetime import datetime
from openpyxl import load_workbook

def save_pareto_set_to_excel(xlsx_path: str, params: dict, pareto_set: list,runtime_sec: float, pubkey, share0, share1, mu, n, n_sq):
    """将非支配解集保存到 Excel 文件"""
    xlsx_path = ensure_results_sheet(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb["Results"]

    # 遍历所有 Pareto 解
    for sol_idx in pareto_set:
        x, enc_cost, enc_qual = sol_idx

        # 解密成本和质量
        dec_cost = combine_shares(
            partial_decrypt(enc_cost, share0, pubkey, n_sq),
            partial_decrypt(enc_cost, share1, pubkey, n_sq),
            mu, n
        )
        dec_qual = combine_shares(
            partial_decrypt(enc_qual, share0, pubkey, n_sq),
            partial_decrypt(enc_qual, share1, pubkey, n_sq),
            mu, n
        )

        # 保存每个解的信息
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            int(params["num_tasks"]), int(params["num_workers"]), int(params["min_assign"]),
            int(params["num_iter"]), int(params["pop_size"]),
            int(dec_cost), int(dec_qual), len(pareto_set),  # pareto_count 记录解集的大小
            float(f"{runtime_sec:.2f}"),  # runtime_sec 留空，后续可以根据需要填充
            json.dumps(x, ensure_ascii=False)
        ])
    wb.save(xlsx_path)

def ensure_results_sheet(xlsx_path: str):
    """确保 xlsx 存在且含 Results 表（带表头）；不存在就创建。返回最终可写入的路径。"""
    from openpyxl import Workbook, load_workbook

    # 若目录不存在先建目录
    dir_ = os.path.dirname(os.path.abspath(xlsx_path))
    if dir_ and not os.path.exists(dir_):
        os.makedirs(dir_, exist_ok=True)

    if not os.path.exists(xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append([
            "timestamp","num_tasks","num_workers","min_assign","num_iter","pop_size",
            "best_cost","best_qual","pareto_count","runtime_sec","best_x_json"
        ])
        wb.save(xlsx_path)
        return xlsx_path

    # 已存在则确保有 Results 表和表头
    wb = load_workbook(xlsx_path)
    if "Results" not in wb.sheetnames:
        ws = wb.create_sheet("Results")
        ws.append([
            "timestamp","num_tasks","num_workers","min_assign","num_iter","pop_size",
            "best_cost","best_qual","pareto_count","runtime_sec","best_x_json"
        ])
        wb.save(xlsx_path)
    return xlsx_path

def append_run_result(xlsx_path: str, params: dict, best_cost: int, best_qual: int,
                      pareto_count: int, runtime_sec: float, best_x):
    """将一次运行结果追加到 Results 表"""
    from openpyxl import load_workbook
    xlsx_path = ensure_results_sheet(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb["Results"]
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        int(params["num_tasks"]), int(params["num_workers"]), int(params["min_assign"]),
        int(params["num_iter"]), int(params["pop_size"]),
        int(best_cost), int(best_qual), int(pareto_count), float(f"{runtime_sec:.2f}"),
        json.dumps(best_x, ensure_ascii=False)
    ])
    wb.save(xlsx_path)

# -----------------------------
# 主运行流程
# -----------------------------
def main():
    # ✅ 参数设置（从 Excel 读取多组；每行一组参数）
    import pandas as pd

    excel_path = 'data_nsga.xlsx'  # 你的参数表；Sheet 名：Config
    df_cfg = pd.read_excel(excel_path, sheet_name='Config')

    # 要求表头包含以下列名（不区分大小写也可自己统一转小写）：
    # num_tasks, num_workers, min_assign, num_iter, pop_size
    need = {'num_tasks', 'num_workers', 'min_assign', 'num_iter', 'pop_size'}
    lowermap = {c.lower(): c for c in df_cfg.columns}
    if not need.issubset(lowermap.keys()):
        raise ValueError(f"Config表头缺少列：{need - set(lowermap.keys())}，请确保含有 {need}")

    # 循环每一行参数组合：对每一组都完整执行下面的原始流程
    for _row_i, row in df_cfg.iterrows():

        print(f"\n==================== 参数组合 #{_row_i + 1} ====================")
        num_tasks = int(row[lowermap['num_tasks']])
        num_workers = int(row[lowermap['num_workers']])
        min_assign = int(row[lowermap['min_assign']])
        num_iter = int(row[lowermap['num_iter']])
        pop_size = int(row[lowermap['pop_size']])

        print(f"本次运行参数：num_tasks={num_tasks}, num_workers={num_workers}, "
              f"min_assign={min_assign}, num_iter={num_iter}, pop_size={pop_size}")
        # 执行每组参数 5 次
        for run in range(5):  # 内部循环，执行 5 次
            print(f"执行第 {run + 1} 次...")
            # 清空旧文件
            for file in ["enc_worker_data.pkl", "threshold_key_shares.pkl", "offline_cache.pkl"]:
                if os.path.exists(file):
                    os.remove(file)

            print("🔐 加密众包优化系统启动...")

            # ✅ 上传加密数据（含位置与最大距离）
            simulate_worker_upload(num_tasks=num_tasks, num_workers=num_workers)

            with open('enc_worker_data.pkl', 'rb') as f:
                data = pickle.load(f)
            with open('threshold_key_shares.pkl', 'rb') as f:
                key_parts = pickle.load(f)

            (pubkey, enc_costs, enc_quals, enc_weights,
             enc_task_locs, enc_worker_locs, enc_max_dists,
             raw_costs, raw_quals, raw_weights,
             raw_task_locs, raw_worker_locs, raw_max_dists,
             share0, share1, mu, n, n_sq, max_budgets) = prepare_data(num_tasks=num_tasks, num_workers=num_workers)

            print("\n📌 工人成本权重向量:")
            print(raw_weights)

            print("\n📌 工人成本矩阵:")
            for i in range(num_workers):
                worker_costs = [raw_costs[t][i] for t in range(num_tasks)]
                print(f"工人 {i + 1}: {worker_costs}")

            print("\n📌 工人位置与最大可接受距离:")
            for i, ((x, y), dmax) in enumerate(zip(raw_worker_locs, raw_max_dists)):
                print(f"工人 {i + 1}: ({x}, {y})，最大距离: {dmax}")

            print("\n📌 任务质量矩阵:")
            for t in range(num_tasks):
                print(f"任务 {t+1}: {raw_quals[t]}")

            # ✅ 打印明文位置信息
            print("\n📌 任务位置坐标:")
            for t, (x, y) in enumerate(raw_task_locs):
                print(f"任务 {t + 1}: ({x}, {y})")

            # ✅ 离线缓存准备
            if not os.path.exists("offline_cache.pkl"):
                generate_offline_cache(pubkey, num_sets=300, num_workers=16)
            print(f"✅ 已加载离线缓存")

            # ✅ 生成 reach[t][i] 矩阵
            print("\n🧮 正在计算可达性矩阵 reach[t][i]...")
            reach = check_reachability_matrix(enc_task_locs, enc_worker_locs, enc_max_dists,
                                              pubkey, share0, share1, mu, n, n_sq)
            print("\n📌 可达性矩阵（1 = 可分配，0 = 超距）:")
            for t in range(num_tasks):
                print(f"任务 {t+1}: {reach[t]}")

            check_task_reachability(reach)

            # ✅ 初始化种群（考虑可达性）
            population = [generate_random_matrix_with_reach(num_tasks, num_workers, min_assign, reach)
                          for _ in range(pop_size)]

            # ✅ 启动优化
            start_time = time.time()
            # 替换为
            best_x, enc_best_cost, enc_best_qual, obj_list = run_nsga2(
                pubkey=pubkey,
                enc_costs=enc_costs,
                enc_quals=enc_quals,
                enc_weights=enc_weights,
                share0=share0,
                share1=share1,
                mu=mu,
                n=n,
                n_sq=n_sq,
                reach=reach,
                num_iter=num_iter,
                pop_size=pop_size,
                num_tasks=num_tasks,
                min_assign=min_assign,
                fast_mode=False,
                init_population=population,
                max_budgets=max_budgets  # 传递最大预算
            )

            print("当前 obj_list 内容：")
            for i, (x, enc_cost, enc_qual) in enumerate(obj_list):
                dec_cost = combine_shares(
                    partial_decrypt(enc_cost, share0, pubkey, n_sq),
                    partial_decrypt(enc_cost, share1, pubkey, n_sq),
                    mu, n
                )
                dec_qual = combine_shares(
                    partial_decrypt(enc_qual, share0, pubkey, n_sq),
                    partial_decrypt(enc_qual, share1, pubkey, n_sq),
                    mu, n
                )
                print(f"解 {i}: 任务分配矩阵 = {x}, 成本 = {dec_cost}, 质量 = {dec_qual}")

            pareto_indices = fast_non_dominated_sort_enc(obj_list, pubkey, share0, share1, mu, n, n_sq)[0]
            print(f"\n🎯 非支配解集个数：{len(pareto_indices)}")
            for i in pareto_indices:
                x, enc_cost, enc_qual = obj_list[i]
                dec_cost = combine_shares(
                    partial_decrypt(enc_cost, share0, pubkey, n_sq),
                    partial_decrypt(enc_cost, share1, pubkey, n_sq),
                    mu, n
                )
                dec_qual = combine_shares(
                    partial_decrypt(enc_qual, share0, pubkey, n_sq),
                    partial_decrypt(enc_qual, share1, pubkey, n_sq),
                    mu, n
                )
                print(f"🧬 解 {i}: 成本 = {dec_cost}, 质量 = {dec_qual}")

            # ✅ 解密输出结果
            print(f"\n✅ 最终任务分配方案：")
            for t, task_assign in enumerate(best_x):
                print(f"任务 {t + 1}: {task_assign}")

            dec_best_cost = combine_shares(
                partial_decrypt(enc_best_cost, share0, pubkey, n_sq),
                partial_decrypt(enc_best_cost, share1, pubkey, n_sq),
                mu, n
            )
            dec_best_qual = combine_shares(
                partial_decrypt(enc_best_qual, share0, pubkey, n_sq),
                partial_decrypt(enc_best_qual, share1, pubkey, n_sq),
                mu, n
            )
            print(f"\n🔓 解密后目标值：")
            print(f"📉 总成本 = {dec_best_cost}")
            print(f"📈 总质量 = {dec_best_qual}")
            print(f"\n⏱️ 本次运行耗时：{time.time() - start_time:.2f} 秒")
            # ===== 你的原始逻辑到此结束；下一轮循环自动开始 =====

            # ===== 写入 Results 工作表 =====
            excel_path = "data_nsga.xlsx"  # 自己的结果文件名/路径
            params_for_log = {
                "num_tasks": num_tasks,
                "num_workers": num_workers,
                "min_assign": min_assign,
                "num_iter": num_iter,
                "pop_size": pop_size
            }
            runtime = time.time() - start_time
            pareto_indices = fast_non_dominated_sort_enc(obj_list, pubkey, share0, share1, mu, n, n_sq)[0]

            # 保存所有非支配解集到 Excel
            print(f"✅ 正在保存非支配解集到 Excel ...")
            save_pareto_set_to_excel(
                xlsx_path=excel_path,
                params=params_for_log,
                pareto_set=[obj_list[i] for i in pareto_indices],
                runtime_sec=runtime,
                pubkey=pubkey,
                share0=share0,
                share1=share1,
                mu=mu,
                n=n,
                n_sq=n_sq
            )

            print(f"✅ 结果已保存到 {excel_path} -> Results")

# ✅ 程序入口
if __name__ == "__main__":
    main()