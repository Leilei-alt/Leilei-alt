# -*- coding: utf-8 -*-
import random
import numpy as np
import pickle
import os
import math
import time
from functools import partial
from concurrent.futures import ProcessPoolExecutor
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm

# ===== FastPai 适配层（替代 phe）========================================
import math, random, secrets
from dataclasses import dataclass
try:
    from Crypto.Util.number import getPrime, isPrime
except Exception:
    raise RuntimeError("需要安装 pycryptodome：pip install pycryptodome")

def _gen_PQ_like(p_bits, pprime_bits):
    """
    生成 P = 2*p*p' + 1 为素数；返回 (P, p, p')
    """
    while True:
        p  = getPrime(p_bits)
        p_ = getPrime(pprime_bits) | 1  # 确保奇数
        P  = 2*p*p_ + 1
        if isPrime(P):
            return P, p, p_

@dataclass
class FastPaiPublicKey:
    n: int
    h: int
    scale: int = 10**6  # 若加密 float，可用定点；你当前不需要解密 crowding，可忽略

    def __post_init__(self):
        self.n_sq   = self.n * self.n
        # 给 SMUL/缓存使用的明文上界，用个保守值即可
        self.max_int = self.n // 3

    def encrypt(self, m):
        # 与 phe 接口保持一致：支持 int（必要时也允许 float→定点）
        if isinstance(m, float):
            m = int(round(m * self.scale))
        else:
            m = int(m)
        c1 = pow(1 + self.n, m, self.n_sq)
        r  = secrets.randbits(256)
        c2 = pow(pow(self.h, r, self.n), self.n, self.n_sq)  # (h^r)^N mod N^2
        return FastPaiEncrypted((c1 * c2) % self.n_sq, self)

class FastPaiEncrypted:
    def __init__(self, c, public_key: FastPaiPublicKey):
        self.c = int(c)
        self.public_key = public_key

    # 兼容你代码里的 ciphertext(False)
    def ciphertext(self, raw=False):
        return self.c

    # 密文加：Enc(a) * Enc(b) = Enc(a+b)
    def __add__(self, other):
        if isinstance(other, FastPaiEncrypted):
            return FastPaiEncrypted((self.c * other.c) % self.public_key.n_sq, self.public_key)
        elif isinstance(other, int):
            c_plain = pow(1 + self.public_key.n, other, self.public_key.n_sq)
            return FastPaiEncrypted((self.c * c_plain) % self.public_key.n_sq, self.public_key)
        else:
            raise TypeError(f"Unsupported add with {type(other)}")
    __radd__ = __add__

    # 密文减：Enc(a) * Enc(b)^{-1} = Enc(a-b)
    def __sub__(self, other):
        if isinstance(other, FastPaiEncrypted):
            inv = pow(other.c, -1, self.public_key.n_sq)
            return FastPaiEncrypted((self.c * inv) % self.public_key.n_sq, self.public_key)
        elif isinstance(other, int):
            return self + (-int(other))
        else:
            raise TypeError(f"Unsupported sub with {type(other)}")

    # 标量乘（plaintext k）：Enc(a)^k = Enc(k*a)；k 只能是整数
    def __mul__(self, k):
        if not isinstance(k, int):
            raise TypeError("Only integer scalar supported for ciphertext * k")
        if k == 0:
            # Enc(0) 的一种合法表示（随机性固定为1）
            return FastPaiEncrypted(1, self.public_key)
        if k > 0:
            return FastPaiEncrypted(pow(self.c, k, self.public_key.n_sq), self.public_key)
        # k < 0 ：取逆即可
        pos = pow(self.c, -k, self.public_key.n_sq)
        inv = pow(pos, -1, self.public_key.n_sq)
        return FastPaiEncrypted(inv, self.public_key)
    __rmul__ = __mul__

def fastpai_generate_threshold_keypair(kappa=112):
    """
    生成 FastPai 公钥/阈值私钥分片
    - P=2pp'+1, Q=2qq'+1 为素数
    - N=PQ, alpha=pq, beta=p'q'
    - h = -y^{2beta} (mod N)
    - 私钥指数=2*alpha，并做两份加法分片：share0+share1=2*alpha
    - 为了少改你现有 combine_shares：把 inv_2alpha 放到返回字典的 'mu' 字段
    """
    # 典型设置：让 N≈2048 位。可按需调整位长分配。
    p_bits, pprime_bits = 256, 512
    P, p, p_ = _gen_PQ_like(p_bits, pprime_bits)
    Q, q, q_ = _gen_PQ_like(p_bits, pprime_bits)

    N = P * Q
    alpha = p * q
    beta  = p_ * q_

    # 选 y∈Z*_N
    while True:
        y = random.randrange(2, N - 1)
        if math.gcd(y, N) == 1:
            break
    h = (-pow(y, 2 * beta, N)) % N

    two_alpha = 2 * alpha
    inv_2alpha = pow(two_alpha, -1, N)

    share0 = random.randrange(1, two_alpha)
    share1 = two_alpha - share0   # 保证 share0+share1=2α

    pubkey = FastPaiPublicKey(N, h)

    return {
        'pubkey': pubkey,
        'privkey': None,      # FastPai 不需要暴露 p,q
        'share0': share0,
        'share1': share1,
        'lambda': None,       # 兼容旧字段，占位
        'mu': inv_2alpha,     # ⚠️ 注意：这里存的是 inv_2alpha
        'n': N,
        'n_sq': N * N
    }
# ======================================================================

# 原：def generate_threshold_keypair(): ...（整块删掉）
# 新：直接把 FastPai 的生成器映射为旧名字，调用方不需要改
generate_threshold_keypair = fastpai_generate_threshold_keypair

# ---------------------------
# 阈值密钥生成（share0 + share1 = λ）
# ---------------------------
def lcm(a, b):
    return abs(a * b) // math.gcd(a, b)

# ---------------------------
# 加密函数
# ---------------------------
def encrypt(pubkey, m):
    return pubkey.encrypt(m)

# ---------------------------
# 部分解密函数
# ---------------------------
def partial_decrypt(ciphertext_obj, share_i, pubkey, n_sq):
    c = ciphertext_obj.ciphertext(False)  # 获取原始密文整数
    u = pow(c, share_i, n_sq)
    return u  # 返回部分解密结果u

# ---------------------------
# 聚合两份部分解密 → 明文
# ---------------------------
def combine_shares(u0, u1, mu, n, max_bits=4096):
    u = (u0 * u1) % (n * n)  # 合并部分解密结果
    L = (u - 1) // n         # 计算L函数
    m = (L * mu) % n         # 恢复明文
    return int(m)

def to_signed_mod(m: int, n: int) -> int:
    """把模 n 的无符号代表元映射回有符号整数区间 (-n/2, n/2]"""
    return m - n if m > (n // 2) else m


# -------------------------------
# 随机解生成与变异（初始化考虑 reach 约束）
# -------------------------------
def generate_random_matrix_with_reach(num_tasks, num_workers, min_assign, reach):
    """
    - 每个任务分配不少于 min_assign 个工人；
    - 每个工人最多参与 1 个任务；
    - 仅在 reach[t][i] == 1 的条件下才允许分配。
    """
    x = [[0 for _ in range(num_workers)] for _ in range(num_tasks)]
    assigned_workers = set()

    # 先满足每个任务的最小分配
    for t in range(num_tasks):
        candidates = [i for i in range(num_workers) if reach[t][i] == 1 and i not in assigned_workers]
        random.shuffle(candidates)
        for i in candidates[:min_assign]:
            x[t][i] = 1
            assigned_workers.add(i)

    # 剩余工人随机分配至可达任务（仍保持“一人一任务”）
    for i in range(num_workers):
        if i in assigned_workers:
            continue
        available_tasks = [t for t in range(num_tasks) if reach[t][i] == 1]
        if available_tasks:
            t = random.choice(available_tasks)
            x[t][i] = 1
            assigned_workers.add(i)

    return x

# ----------------------------------------------------
# 非支配排序 & 判重（保持不变）
# ----------------------------------------------------
def is_duplicate(new_sol, pareto_set, pubkey, share0, share1, mu, n, n_sq, epsilon_int: int = 0):
    """
    判重：将解密结果做有符号映射后，用整数比较。
    epsilon_int 为整数阈值（0 表示严格相等才算重复）
    """
    # 新解解密 + 映射
    enc_cost_new = new_sol[1]
    enc_qual_new = new_sol[2]
    dec_cost_new = combine_shares(
        partial_decrypt(enc_cost_new, share0, pubkey, n_sq),
        partial_decrypt(enc_cost_new, share1, pubkey, n_sq),
        mu, n
    )
    dec_qual_new = combine_shares(
        partial_decrypt(enc_qual_new, share0, pubkey, n_sq),
        partial_decrypt(enc_qual_new, share1, pubkey, n_sq),
        mu, n
    )
    cost_new = to_signed_mod(dec_cost_new, n)
    qual_new = to_signed_mod(dec_qual_new, n)

    for sol in pareto_set:
        enc_cost_sol = sol[1]
        enc_qual_sol = sol[2]

        dec_cost_sol = combine_shares(
            partial_decrypt(enc_cost_sol, share0, pubkey, n_sq),
            partial_decrypt(enc_cost_sol, share1, pubkey, n_sq),
            mu, n
        )
        dec_qual_sol = combine_shares(
            partial_decrypt(enc_qual_sol, share0, pubkey, n_sq),
            partial_decrypt(enc_qual_sol, share1, pubkey, n_sq),
            mu, n
        )
        cost_sol = to_signed_mod(dec_cost_sol, n)
        qual_sol = to_signed_mod(dec_qual_sol, n)

        # 用整数的 L1 距离（也可以用等值比较）
        l1 = abs(cost_new - cost_sol) + abs(qual_new - qual_sol)
        if l1 <= epsilon_int:
            return True
    return False

def fast_non_dominated_sort_enc(obj_list, pubkey, share0, share1, mu, n, n_sq):
    S = {}
    n_dom = {}
    fronts = [[]]
    for p in range(len(obj_list)):
        S[p] = []
        n_dom[p] = 0
        for q in range(len(obj_list)):
            if p == q:
                continue
            if dominates_enc(obj_list[p], obj_list[q], pubkey, share0, share1, mu, n, n_sq):
                S[p].append(q)
            elif dominates_enc(obj_list[q], obj_list[p], pubkey, share0, share1, mu, n, n_sq):
                n_dom[p] += 1
        if n_dom[p] == 0:
            fronts[0].append(p)

    i = 0
    while len(fronts[i]) > 0:
        next_front = []
        for p in fronts[i]:
            for q in S[p]:
                n_dom[q] -= 1
                if n_dom[q] == 0:
                    next_front.append(q)
        i += 1
        fronts.append(next_front)

    if len(fronts[-1]) == 0:
        fronts.pop()

    # ⚠️ 兜底：如果没有任何个体进入第一前沿，退化为“所有个体同属第一前沿”
    if not fronts and len(obj_list) > 0:
        fronts = [list(range(len(obj_list)))]



    # 去重（按前面判重规则）
    unique_fronts = []
    for front in fronts:
        unique_front = []
        for idx in front:
            if not is_duplicate(
                obj_list[idx],
                [obj_list[i] for i in unique_front],
                pubkey, share0, share1, mu, n, n_sq
            ):
                unique_front.append(idx)
        unique_fronts.append(unique_front)

    return unique_fronts

# ----------------------------------------------------
# 拥塞距离（密文加权代理版本，保持不变）
# ----------------------------------------------------
def compute_weighted_crowding_distance(obj_list, population, front_indices,
                                       pubkey, share0, share1, mu, n, n_sq,
                                       alpha=0.5):
    """
    EncCrowding ≈ α × (enc_max_cost - enc_cost) + (1 - α) × enc_qual
    """
    enc_costs = [obj_list[i][1] for i in front_indices]
    enc_quals = [obj_list[i][2] for i in front_indices]

    max_cost_idx = select_argmax_enc(list(zip(front_indices, enc_costs)),
                                     pubkey, share0, share1, mu, n, n_sq)
    enc_max_cost = obj_list[max_cost_idx][1]

    crowding_dict = {}
    S = 10 ** 6
    A = int(round(alpha * S))
    B = int(round((1 - alpha) * S))

    for i in front_indices:
        enc_cost = obj_list[i][1]
        enc_qual = obj_list[i][2]
        enc_diff = enc_max_cost - enc_cost
        enc_part1 = enc_diff * A
        enc_part2 = enc_qual * B
        enc_crowding = enc_part1 + enc_part2
        crowding_dict[i] = enc_crowding

    return crowding_dict

# ----------------------------------------------------
# Elite/TopK 选择（保持不变）
# ----------------------------------------------------
def integrate_elite_selection(population, crowding_dict, k, pubkey, share0, share1, mu, n, n_sq):
    elite_indices = select_topk_by_enc_value(crowding_dict, k, pubkey, share0, share1, mu, n, n_sq)
    if len(elite_indices) < k:
        supplement = select_random_if_empty(crowding_dict, k - len(elite_indices))
        elite_indices += supplement
    return elite_indices

from fractions import Fraction

def select_best_by_custom_score(obj_list, share0, share1, pubkey, mu, n, n_sq,
                                alpha=0.7, beta=0.3, denom_eps_int: int = 1):
    """
    用纯整数进行比较，避免超大 int 转 float 溢出。
    评分：score = (alpha * qual) / (beta * cost)
    比较两个解 a、b 的分数，用交叉相乘：
      (a_num / a_den) ? (b_num / b_den)
    其中：
      a_num = alpha_num * beta_den * qual_a
      a_den = beta_num  * alpha_den * max(cost_a, denom_eps_int)   # 防 0/负
    alpha_num/alpha_den、beta_num/beta_den 来自 Fraction(alpha), Fraction(beta)
    """
    # 把 alpha、beta 表示成分数，避免浮点
    frac_a = Fraction(alpha).limit_denominator(10**6)  # 0.7 -> 7/10
    frac_b = Fraction(beta ).limit_denominator(10**6)  # 0.3 -> 3/10

    aN, aD = frac_a.numerator, frac_a.denominator
    bN, bD = frac_b.numerator, frac_b.denominator

    best_idx = None
    best_num = None
    best_den = None

    for i, (_, enc_cost, enc_qual) in enumerate(obj_list):
        # 解密并做有符号映射
        dec_cost = combine_shares(
            partial_decrypt(enc_cost, share0, pubkey, n_sq),
            partial_decrypt(enc_cost, share1, pubkey, n_sq),
            mu, n
        )
        dec_qual = combine_shares(
            partial_decrypt(enc_qual, share0, pubkey, n_sq),
            partial_decrypt(enc_qual, share1, pubkey, n_sq),
            mu, n
        )
        cost = to_signed_mod(dec_cost, n)
        qual = to_signed_mod(dec_qual, n)

        # 分母防 0/负（按你的任务定义 cost 应该为正，但加个兜底更稳）
        denom_cost = cost if cost > 0 else denom_eps_int

        # 计算“整数化”的分子和分母（均为正整型）
        num = aN * bD * int(qual)       # alpha_num * beta_den * qual
        den = bN * aD * int(denom_cost) # beta_num  * alpha_den * cost

        # 第一个候选或更优就更新（交叉相乘避免浮点）
        if best_idx is None or (num * best_den > best_num * den):
            best_idx = i
            best_num = num
            best_den = den

    return best_idx




# ================== 密文乘法（SMUL）— 与你现有缓存/阈值流程兼容 ==================
def secure_multiply(enc_x, enc_y, pubkey, share0, share1, mu, n, n_sq):
    """
    计算 Enc(x) * Enc(y) ≡ Enc(x·y) 的协议化实现（双服务阈值、带离线缓存）。
    依赖：
      - get_next_cache_process_safe(pubkey) 产生/加载一组 (r1, r2, Enc(r1), Enc(r2), Enc(r1*r2), Enc(0), Enc(1))
      - partial_decrypt / combine_shares 做局部+合并解密
      - FastPaiEncrypted 支持与整数的标量乘（__mul__）
    """
    cache = get_next_cache_process_safe(pubkey=pubkey)

    r1 = cache['r1']
    r2 = cache['r2']
    enc_r1 = cache['enc_r1']
    enc_r2 = cache['enc_r2']
    enc_r1r2 = cache['enc_r1r2']

    # 加噪：x' = x + r1,  y' = y + r2
    enc_x_ = enc_x + enc_r1
    enc_y_ = enc_y + enc_r2

    # 解密 y' 得到明文 (y + r2)
    u0 = partial_decrypt(enc_y_, share0, pubkey, n_sq)
    u1 = partial_decrypt(enc_y_, share1, pubkey, n_sq)
    y_plus_r2 = combine_shares(u0, u1, mu, n)  # 明文整数

    # 组合：Enc(x·(y+r2)) - Enc(x·r2) - Enc(y·r1) - Enc(r1·r2)
    # 注意：Enc(a)*k 是“标量乘”（幂运算同态），k 必须是整数
    enc_xy_noisy = enc_x_ * int(y_plus_r2)      # Enc(x + r1) * (y + r2)
    enc_x_r2     = enc_x * int(r2) * -1         # - Enc(x·r2)
    enc_y_r1     = enc_y * int(r1) * -1         # - Enc(y·r1)
    enc_r1r2_neg = enc_r1r2 * -1                # - Enc(r1·r2)

    result = enc_xy_noisy + enc_x_r2 + enc_y_r1 + enc_r1r2_neg
    return result
# =====================================================================

# -------------------------------
# 密文目标函数评估（保持不变）
# -------------------------------
def evaluate_cost_stable_smulg(x, enc_costs, enc_weights, pubkey, share0, share1, mu, n, n_sq):
    """
    现在的成本 = ∑ Enc(cost_{t,i})，不再与 enc_weights 相乘
    （保留 enc_weights 形参是为了不改调用处签名，但这里不用它）
    """
    T = len(x)
    N = len(x[0])

    total_cost = None
    for t in range(T):
        for i in range(N):
            if x[t][i] == 1:
                enc_c = enc_costs[t][i]             # Enc(cost_{t,i})
                total_cost = enc_c if total_cost is None else (total_cost + enc_c)
    return total_cost if total_cost else pubkey.encrypt(0)


def evaluate_quality_stable(x, enc_quals, enc_weights,
                            pubkey, share0, share1, mu, n, n_sq):
    """
    现在的质量 = ∑ Enc(quality_{t,i} * weight_i)
    使用 secure_multiply 做密文乘法，再同态加法相加
    """
    T = len(x)
    N = len(x[0])

    total_qual = None
    for t in range(T):
        for i in range(N):
            if x[t][i] == 1:
                enc_q = enc_quals[t][i]              # Enc(quality_{t,i})
                enc_w = enc_weights[i]               # Enc(weight_i)
                enc_qw = secure_multiply(enc_q, enc_w, pubkey, share0, share1, mu, n, n_sq)
                total_qual = enc_qw if total_qual is None else (total_qual + enc_qw)
    return total_qual if total_qual else pubkey.encrypt(0)


def evaluate_individual_parallel(x, enc_costs, enc_quals, enc_weights,
                                 pubkey, share0, share1, mu, n, n_sq,
                                 min_assign, fast_mode=False, max_budgets=None):
    T = len(x); N = len(x[0])
    valid = True

    # 预算与 min_assign 检查（解密累加成本）
    if max_budgets is not None:
        for t in range(T):
            dec_task_cost = 0
            for i in range(N):
                if x[t][i] == 1:
                    dec_task_cost += combine_shares(
                        partial_decrypt(enc_costs[t][i], share0, pubkey, n_sq),
                        partial_decrypt(enc_costs[t][i], share1, pubkey, n_sq),
                        mu, n
                    )
            if dec_task_cost > max_budgets[t]:
                valid = False
                break

    for t in range(T):
        if sum(x[t]) < min_assign:
            valid = False
            break

    if fast_mode:
        cost = sum(i * x[t][i] for t in range(T) for i in range(N))
        qual = sum(10 * x[t][i] for t in range(T) for i in range(N))
        return {"x": x, "enc_cost": pubkey.encrypt(cost), "enc_qual": pubkey.encrypt(qual), "key": str(x)}

    if not valid:
        enc_cost = pubkey.encrypt(999999)
        enc_qual = pubkey.encrypt(0)
    else:
        enc_cost = evaluate_cost_stable_smulg(x, enc_costs, enc_weights, pubkey, share0, share1, mu, n, n_sq)
        # 现在（加权质量）：
        enc_qual = evaluate_quality_stable(
            x, enc_quals, enc_weights,
            pubkey, share0, share1, mu, n, n_sq
        )

    return {"x": x, "enc_cost": enc_cost, "enc_qual": enc_qual, "key": str(x)}

def parallel_evaluate_population(population, enc_costs, enc_quals, enc_weights,
                                 pubkey, share0, share1, mu, n, n_sq,
                                 min_assign, fast_mode=False,
                                 use_cache=True, max_budgets=None):
    from concurrent.futures import ProcessPoolExecutor, as_completed
    from tqdm import tqdm
    import multiprocessing

    global decrypt_cache
    if 'decrypt_cache' not in globals():
        decrypt_cache = {}

    results = []
    with ProcessPoolExecutor(max_workers=multiprocessing.cpu_count()) as executor:
        futures = []
        for x in population:
            key = str(x)
            if use_cache and key in decrypt_cache:
                results.append(decrypt_cache[key])
            else:
                futures.append(executor.submit(
                    evaluate_individual_parallel, x,
                    enc_costs, enc_quals, enc_weights,
                    pubkey, share0, share1, mu, n, n_sq,
                    min_assign, fast_mode, max_budgets
                ))

        for f in tqdm(as_completed(futures), total=len(futures), desc="⚡并行评估中"):
            result = f.result()
            results.append((result["x"], result["enc_cost"], result["enc_qual"]))
            if use_cache:
                decrypt_cache[result["key"]] = (result["x"], result["enc_cost"], result["enc_qual"])

    return results

# ----------------------------------------------------
# 支配关系与密文选择（保持不变）
# ----------------------------------------------------
def dominates_enc(ind1, ind2, pubkey, share0, share1, mu, n, n_sq):
    enc_cost1, enc_qual1 = ind1[1], ind1[2]
    enc_cost2, enc_qual2 = ind2[1], ind2[2]

    # cost1 < cost2 以及 qual1 > qual2 都用严格比较
    comp1 = secure_compare(enc_cost1, enc_cost2, pubkey, share0, share1, mu, n, n_sq, strict=True)
    comp2 = secure_compare(enc_qual2, enc_qual1, pubkey, share0, share1, mu, n, n_sq, strict=True)

    u01 = partial_decrypt(comp1, share0, pubkey, n_sq)
    u02 = partial_decrypt(comp2, share0, pubkey, n_sq)
    u11 = partial_decrypt(comp1, share1, pubkey, n_sq)
    u12 = partial_decrypt(comp2, share1, pubkey, n_sq)

    b1 = combine_shares(u01, u11, mu, n)  # cost1 < cost2 ?
    b2 = combine_shares(u02, u12, mu, n)  # qual1 > qual2 ?

    return b1 == 1 and b2 == 1


def select_topk_by_enc_value(enc_dict, k, pubkey, share0, share1, mu, n, n_sq):
    selected = []
    remaining = list(enc_dict.keys())
    while len(selected) < k and remaining:
        max_idx = remaining[0]
        for i in remaining[1:]:
            comp = secure_compare(enc_dict[i], enc_dict[max_idx], pubkey, share0, share1, mu, n, n_sq)
            u0 = partial_decrypt(comp, share0, pubkey, n_sq)
            u1 = partial_decrypt(comp, share1, pubkey, n_sq)
            result = combine_shares(u0, u1, mu, n)
            if result == 1:
                max_idx = i
        selected.append(max_idx)
        remaining.remove(max_idx)
    return selected

def select_argmax_enc(index_value_pairs, pubkey, share0, share1, mu, n, n_sq):
    max_index, max_enc = index_value_pairs[0]
    for idx, enc in index_value_pairs[1:]:
        enc_cmp = secure_compare(enc, max_enc, pubkey, share0, share1, mu, n, n_sq)
        u0 = partial_decrypt(enc_cmp, share0, pubkey, n_sq)
        u1 = partial_decrypt(enc_cmp, share1, pubkey, n_sq)
        bit = combine_shares(u0, u1, mu, n)
        if bit == 1:
            max_index, max_enc = idx, enc
    return max_index

def select_random_if_empty(enc_dict, k):
    all_indices = list(enc_dict.keys())
    if not all_indices:
        return []
    return random.sample(all_indices, min(k, len(all_indices)))

# ----------------------------------------------------
# 距离/可达性（保持不变）
# ----------------------------------------------------
def secure_compare(enc_x, enc_y, pubkey, share0, share1, mu, n, n_sq, strict=False):
    """
    返回 Enc(1) 若 (enc_y ? enc_x) 成立，否则返回 Enc(0)
      - strict=False : 判断 y >= x
      - strict=True  : 判断 y >  x
    """
    cache = get_next_cache_process_safe(pubkey=pubkey)
    r = cache['r1']
    r_dash = random.randint(n // 5, n // 4)

    # 基础差值：strict 时不要 +1（严格），否则 +1（非严格）
    enc_diff_base = enc_y - enc_x
    if not strict:
        enc_diff_base = enc_diff_base + pubkey.encrypt(1)  # y - x + 1  ⇔ y >= x

    enc_r_diff = enc_diff_base * r
    enc_full = enc_r_diff + pubkey.encrypt(r_dash)

    u0 = partial_decrypt(enc_full, share0, pubkey, n_sq)
    u1 = partial_decrypt(enc_full, share1, pubkey, n_sq)
    d = combine_shares(u0, u1, mu, n)

    return cache['enc_1'] if d > r_dash else cache['enc_0']


def secure_manhattan_distance(task_loc_enc, worker_loc_enc, pubkey, share0, share1, mu, n, n_sq):
    enc_xt, enc_yt = task_loc_enc
    enc_xi, enc_yi = worker_loc_enc
    enc_dx = enc_xt - enc_xi
    enc_dy = enc_yt - enc_yi

    def abs_enc(enc_a):
        enc_neg_a = enc_a * -1
        cmp = secure_compare(enc_a, enc_neg_a, pubkey, share0, share1, mu, n, n_sq)
        u0 = partial_decrypt(cmp, share0, pubkey, n_sq)
        u1 = partial_decrypt(cmp, share1, pubkey, n_sq)
        b = combine_shares(u0, u1, mu, n)
        return enc_neg_a if b == 1 else enc_a

    abs_dx = abs_enc(enc_dx)
    abs_dy = abs_enc(enc_dy)
    return abs_dx + abs_dy

def check_reachability_matrix(enc_task_locs, enc_worker_locs, enc_max_dists,
                               pubkey, share0, share1, mu, n, n_sq):
    num_tasks = len(enc_task_locs)
    num_workers = len(enc_worker_locs)
    reach = [[0 for _ in range(num_workers)] for _ in range(num_tasks)]
    for t in range(num_tasks):
        for i in range(num_workers):
            enc_dist = secure_manhattan_distance(enc_task_locs[t], enc_worker_locs[i],
                                                 pubkey, share0, share1, mu, n, n_sq)
            cmp = secure_compare(enc_dist, enc_max_dists[i], pubkey, share0, share1, mu, n, n_sq)
            u0 = partial_decrypt(cmp, share0, pubkey, n_sq)
            u1 = partial_decrypt(cmp, share1, pubkey, n_sq)
            b = combine_shares(u0, u1, mu, n)
            reach[t][i] = 1 if b == 1 else 0
    return reach

def check_task_reachability(reach):
    for t_idx, row in enumerate(reach):
        if sum(row) == 0:
            raise ValueError(f"❌ 错误：任务 {t_idx + 1} 无法被任何工人完成（reach 全为0）")

# ----------------------------------------------------
# 离线缓存（保持不变）
# ----------------------------------------------------
def generate_single_cache(pubkey):
    max_val = pubkey.max_int // 4
    while True:
        r1 = random.randint(1, 10000)
        r2 = random.randint(1, 10000)
        if r1 * r2 <= max_val:
            break
    return {
        'r1': r1,
        'r2': r2,
        'enc_r1': pubkey.encrypt(r1),
        'enc_r2': pubkey.encrypt(r2),
        'enc_r1r2': pubkey.encrypt(r1 * r2),
        'enc_0': pubkey.encrypt(0),
        'enc_1': pubkey.encrypt(1)
    }

def generate_offline_cache(pubkey, num_sets=600, num_workers=16):
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        cache_list = list(executor.map(lambda _: generate_single_cache(pubkey), range(num_sets)))
    with open("offline_cache.pkl", "wb") as f:
        pickle.dump(cache_list, f)
    print(f"✅ 多线程生成完成，共 {num_sets} 组扰动元组，使用 {num_workers} 线程")

def load_offline_cache():
    with open("offline_cache.pkl", "rb") as f:
        return pickle.load(f)

def get_next_cache_process_safe(pubkey=None, refill_num=100):
    if not os.path.exists("offline_cache.pkl"):
        if pubkey is None:
            raise ValueError("❌ 无 offline_cache.pkl 且未提供 pubkey")
        generate_offline_cache(pubkey, num_sets=refill_num)
    with open("offline_cache.pkl", "rb") as f:
        cache_list = pickle.load(f)
    return random.choice(cache_list)

# ----------------------------------------------------
# 固定数据 + 加密上传（保持不变）
# ----------------------------------------------------
def simulate_worker_upload(num_tasks=7, num_workers=20):
    if num_tasks != 7 or num_workers != 20:
        raise ValueError("固定数据仅支持num_tasks=7和num_workers=20")

    keys = generate_threshold_keypair()
    pubkey = keys['pubkey']
    share0, share1 = keys['share0'], keys['share1']
    mu, n, n_sq = keys['mu'], keys['n'], keys['n_sq']

    # --------------------------
    # 固定的工人成本权重向量（20 名工人）
    # --------------------------
    raw_weights = [2, 2, 5, 1, 1, 4, 3, 5, 3, 2, 3, 3, 4, 2, 3, 5, 2, 4, 1, 3]

    # --------------------------
    # 固定的工人成本矩阵 (20 个工人 × 7 个任务)
    # 注意：这里按“工人×任务”填写；下面一行会转置成“任务×工人”使用
    # --------------------------
    raw_costs = [
        [26, 14, 21, 29, 30, 18, 22],  # 工人1
        [15, 29, 11, 11, 19, 17, 24],  # 工人2
        [12, 28, 23, 17, 22, 16, 19],  # 工人3
        [30, 15, 10, 18, 10, 20, 13],  # 工人4
        [14, 21, 15, 19, 26, 12, 18],  # 工人5
        [21, 30, 29, 20, 15, 14, 20],  # 工人6
        [11, 13, 11, 10, 11, 12, 16],  # 工人7
        [22, 10, 10, 28, 23, 19, 21],  # 工人8
        [30, 23, 20, 10, 19, 18, 24],  # 工人9
        [27, 13, 18, 10, 16, 15, 20],  # 工人10
        [20, 13, 11, 20, 26, 14, 17],  # 工人11
        [12, 23, 28, 13, 24, 16, 20],  # 工人12
        [23, 13, 12, 23, 29, 21, 18],  # 工人13
        [27, 24, 23, 28, 30, 22, 25],  # 工人14
        [26, 15, 12, 29, 12, 18, 21],  # 工人15
        [18, 16, 20, 22, 14, 17, 19],  # 工人16
        [19, 21, 17, 15, 18, 13, 20],  # 工人17
        [16, 19, 22, 14, 20, 18, 23],  # 工人18
        [24, 17, 15, 21, 16, 19, 18],  # 工人19
        [17, 20, 19, 16, 22, 20, 21],  # 工人20
    ]
    # 转置成本矩阵（原数据是工人×任务，需要转为任务×工人）
    raw_costs = [[raw_costs[i][t] for i in range(num_workers)] for t in range(num_tasks)]

    # --------------------------
    # 固定的任务质量矩阵 (7 个任务 × 20 个工人)
    # 每行对应一个任务，对应 20 名工人的质量评分（建议范围 5~15）
    # --------------------------
    raw_quals = [
        [13, 8, 12, 7, 8, 9, 7, 15, 9, 10, 6, 15, 8, 6, 8, 11, 12, 9, 7, 10],  # 任务1
        [5, 13, 13, 15, 6, 9, 6, 13, 11, 15, 15, 8, 14, 7, 6, 10, 9, 12, 8, 11],  # 任务2
        [6, 11, 6, 12, 8, 14, 10, 15, 10, 7, 10, 12, 11, 11, 13, 9, 12, 10, 7, 14],  # 任务3
        [6, 9, 13, 13, 10, 11, 6, 5, 6, 13, 8, 8, 14, 5, 8, 12, 11, 9, 10, 7],  # 任务4
        [10, 15, 13, 14, 12, 11, 6, 11, 15, 5, 15, 6, 6, 6, 7, 13, 12, 10, 9, 8],  # 任务5
        [9, 12, 11, 10, 7, 13, 8, 12, 9, 11, 10, 9, 12, 8, 10, 14, 11, 12, 9, 13],  # 任务6
        [8, 10, 12, 9, 11, 12, 9, 13, 10, 9, 11, 12, 10, 12, 9, 11, 10, 13, 8, 12],  # 任务7
    ]

    # --------------------------
    # 固定的工人位置与最大距离（20 名工人）
    # --------------------------
    raw_worker_locs = [
        (7, 30),  # 工人1
        (34, 24),  # 工人2
        (15, 16),  # 工人3
        (24, 36),  # 工人4
        (46, 6),  # 工人5
        (42, 26),  # 工人6
        (8, 6),  # 工人7
        (0, 9),  # 工人8
        (17, 8),  # 工人9
        (37, 32),  # 工人10
        (36, 9),  # 工人11
        (18, 30),  # 工人12
        (45, 20),  # 工人13
        (26, 7),  # 工人14
        (25, 46),  # 工人15
        (12, 22),  # 工人16
        (33, 14),  # 工人17
        (5, 27),  # 工人18
        (28, 18),  # 工人19
        (9, 41),  # 工人20
    ]
    raw_max_dists = [
        41, 35, 47, 47, 27, 35, 22, 39, 24, 27,
        43, 32, 28, 35, 42, 30, 33, 36, 29, 40,
    ]

    # --------------------------
    # 固定的任务位置（7 个任务）
    # --------------------------
    raw_task_locs = [
        (21, 49),  # 任务1
        (43, 13),  # 任务2
        (3, 12),  # 任务3
        (5, 3),  # 任务4
        (23, 39),  # 任务5
        (15, 25),  # 任务6
        (38, 28),  # 任务7
    ]

    enc_weights = [encrypt(pubkey, w) for w in raw_weights]
    enc_costs = [[encrypt(pubkey, raw_costs[t][i]) for i in range(num_workers)] for t in range(num_tasks)]
    enc_quals = [[encrypt(pubkey, raw_quals[t][i]) for i in range(num_workers)] for t in range(num_tasks)]
    enc_task_locs = [(encrypt(pubkey, x), encrypt(pubkey, y)) for x, y in raw_task_locs]
    enc_worker_locs = [(encrypt(pubkey, x), encrypt(pubkey, y)) for x, y in raw_worker_locs]
    enc_max_dists = [encrypt(pubkey, d) for d in raw_max_dists]

    with open("enc_worker_data.pkl", "wb") as f:
        pickle.dump({
            "pubkey": pubkey,
            "costs": enc_costs,
            "quals": enc_quals,
            "raw_costs": raw_costs,
            "raw_quals": raw_quals,
            "raw_weights": raw_weights,
            "enc_weights": enc_weights,
            "raw_task_locs": raw_task_locs,
            "raw_worker_locs": raw_worker_locs,
            "raw_max_dists": raw_max_dists,
            "enc_task_locs": enc_task_locs,
            "enc_worker_locs": enc_worker_locs,
            "enc_max_dists": enc_max_dists
        }, f)

    with open("threshold_key_shares.pkl", "wb") as f:
        pickle.dump({
            "share0": share0,
            "share1": share1,
            "mu": mu,
            "n": n,
            "n_sq": n_sq
        }, f)

    print("✅ 多任务加密数据上传成功（使用固定数据）")

def prepare_data(num_tasks, num_workers):
    simulate_worker_upload(num_tasks=num_tasks, num_workers=num_workers)
    with open('enc_worker_data.pkl', 'rb') as f:
        data = pickle.load(f)
    with open('threshold_key_shares.pkl', 'rb') as f:
        key_parts = pickle.load(f)
    # 定义每个任务的最大预算（7 个任务）
    max_budgets = [100, 150, 200, 120, 180, 200, 190]  # 举例，任务1的预算是100，任务2的预算是150，以此类推
    return (
        data['pubkey'], data['costs'], data['quals'], data['enc_weights'],
        data['enc_task_locs'], data['enc_worker_locs'], data['enc_max_dists'],
        data['raw_costs'], data['raw_quals'], data['raw_weights'],
        data['raw_task_locs'], data['raw_worker_locs'], data['raw_max_dists'],
        key_parts['share0'], key_parts['share1'], key_parts['mu'], key_parts['n'], key_parts['n_sq'],
        max_budgets
    )

# =========================================================
# 🔁🔁 MODE（差分进化）版优化器（新增）
# =========================================================

def sigmoid(x):
    return 1.0 / (1.0 + np.exp(-x))

def binarize_and_repair(mat_float, reach, min_assign):
    """
    将连续矩阵∈R^{T×N} → 二值矩阵，并进行可行性修复：
    - reach[t][i]==0 的位置强制 0
    - 每个工人至多参与 1 个任务（列约束）
    - 每个任务至少分配 min_assign 个工人（行约束）
    """
    T, N = mat_float.shape
    X = (sigmoid(mat_float) >= 0.5).astype(int)

    # reach 约束：不可达位置清零
    reach_np = np.array(reach, dtype=int)
    X = X * reach_np

    # 列约束：每个工人最多 1 个任务 —— 若多于1个，保留得分最高的那个任务
    for i in range(N):
        rows = np.where(X[:, i] == 1)[0]
        if len(rows) > 1:
            # 选择该列在 mat_float 中值最大的行
            best_r = rows[np.argmax(mat_float[rows, i])]
            X[:, i] = 0
            X[best_r, i] = 1

    # 行约束：每个任务至少 min_assign
    for t in range(T):
        cnt = np.sum(X[t, :])
        if cnt < min_assign:
            # 从可达且当前未被占用的工人中，按 mat_float 分数高→低补足
            candidates = [i for i in range(N) if reach[t][i] == 1 and X[:, i].sum() == 0]
            if len(candidates) > 0:
                cand_scores = [(i, mat_float[t, i]) for i in candidates]
                cand_scores.sort(key=lambda z: z[1], reverse=True)
                need = min_assign - cnt
                for i, _ in cand_scores[:need]:
                    X[t, i] = 1
                    # 列约束自动满足，因为只在未被占用的列里选
            # 若仍不足（比如 reach 太稀疏或都被别的任务占用），就从同列里挑分最低的任务释放，再给当前任务
            cnt = np.sum(X[t, :])
            if cnt < min_assign:
                # 释放策略：找那些 reach[t][i]==1 但当前 X[:,i] 已有1 的列，替换给当前任务
                stealable = [i for i in range(N) if reach[t][i]==1 and X[:,i].sum()==1 and X[t,i]==0]
                # 选择“被占用列”中，当前所在任务在该列上的分最低者进行调换
                swap_list = []
                for i in stealable:
                    r_old = int(np.where(X[:, i]==1)[0][0])
                    swap_list.append((i, mat_float[r_old, i], r_old, mat_float[t, i]))
                # 优先换“旧任务分低、当前任务分高”的列
                swap_list.sort(key=lambda z: (z[1]-z[3]))  # 旧-新 越大越差，优先换
                need = min_assign - cnt
                for (i, _, r_old, _) in swap_list[:need]:
                    X[r_old, i] = 0
                    X[t, i] = 1

    return X.tolist()

def make_float_from_binary(population):
    """
    将 0/1 种群转成 float 初始化（0→-2.0，1→+2.0，便于 sigmoid 后明显分开）
    """
    pop_float = []
    for x in population:
        arr = np.array(x, dtype=float)
        arr = np.where(arr>0.5, 2.0, -2.0)
        pop_float.append(arr)
    return pop_float

import numpy as np

import numpy as np

def de_mutation_and_crossover(i, pop_float, F=0.5, CR=0.9, lower=-2.0, upper=2.0, rng=None):
    rng = np.random.default_rng() if rng is None else rng
    pop = np.asarray(pop_float, dtype=float)          # (NP, *G_shape)
    NP = pop.shape[0]

    # 将 i 转为纯 int，避免 numpy 标量带来的比较问题
    if not isinstance(i, (int, np.integer)):
        i = int(np.asarray(i).item())
    if not (0 <= i < NP):
        raise IndexError(f"索引 i={i} 超界（0..{NP-1}）")

    G_shape = pop.shape[1:]                           # 例如 (5, 15)
    D = int(np.prod(G_shape))                         # 例如 75

    # 预先准备边界
    lower_b = np.broadcast_to(lower, G_shape).reshape(D)
    upper_b = np.broadcast_to(upper, G_shape).reshape(D)
    x_i = pop[i].reshape(D)

    # ---------- 人口不足的降级策略（避免直接抛错） ----------
    if NP < 4:
        # 对当前体做小幅高斯扰动，裁剪到边界
        scale = (upper_b - lower_b) * 0.05
        v = np.clip(x_i + rng.normal(0.0, 1.0, size=D) * scale, lower_b, upper_b)
    else:
        # ---------- 标准 DE/rand/1/bin ----------
        pool = np.delete(np.arange(NP), i)            # 去掉 i
        a_idx, b_idx, c_idx = rng.choice(pool, 3, replace=False)
        a = pop[a_idx].reshape(D)
        b = pop[b_idx].reshape(D)
        c = pop[c_idx].reshape(D)
        v = a + F * (b - c)
        v = np.minimum(np.maximum(v, lower_b), upper_b)

    # 交叉：保证至少一位发生交叉
    mask = (rng.random(D) < CR)
    if not mask.any():
        mask[rng.integers(D)] = True

    u = np.where(mask, v, x_i).reshape(G_shape)       # 还原成 (T, N)
    return u



def environment_selection_MODE(parents, parents_objs, offsprings, off_objs,
                               pubkey, share0, share1, mu, n, n_sq, pop_size, alpha_for_crowding=0.5):
    """
    MODE 的环境选择：父代+子代合并→非支配排序→前几层塞满；最后一层用密态加权拥塞距离筛选。
    不足 pop_size 时，从剩余候选中随机补齐（保证人口恒定）。
    """
    combined_pop = parents + offsprings
    combined_obj = parents_objs + off_objs

    fronts = fast_non_dominated_sort_enc(combined_obj, pubkey, share0, share1, mu, n, n_sq)
    new_pop = []
    new_objs = []
    selected_indices = set()

    for f in fronts:
        if len(new_pop) + len(f) <= pop_size:
            for idx in f:
                new_pop.append(combined_pop[idx])
                new_objs.append(combined_obj[idx])
                selected_indices.add(idx)
        else:
            # 需要在该 front 内部挑 K 个
            k = pop_size - len(new_pop)
            if k <= 0:
                break
            crowding_dict = compute_weighted_crowding_distance(
                combined_obj, combined_pop, f,
                pubkey, share0, share1, mu, n, n_sq,
                alpha=alpha_for_crowding
            )
            topk = select_topk_by_enc_value(crowding_dict, k, pubkey, share0, share1, mu, n, n_sq)
            for idx in topk:
                new_pop.append(combined_pop[idx])
                new_objs.append(combined_obj[idx])
                selected_indices.add(idx)
            break

    # --- 兜底：若仍不足 pop_size，则从剩余个体随机补齐 ---
    if len(new_pop) < pop_size:
        all_indices = list(range(len(combined_pop)))
        remaining = [idx for idx in all_indices if idx not in selected_indices]
        need = pop_size - len(new_pop)
        if remaining:
            supplement = random.sample(remaining, min(need, len(remaining)))
            for idx in supplement:
                new_pop.append(combined_pop[idx])
                new_objs.append(combined_obj[idx])

    return new_pop, new_objs




def run_mode(pubkey, enc_costs, enc_quals, enc_weights,
             share0, share1, mu, n, n_sq, reach,
             num_iter=3, pop_size=5,
             num_tasks=3, min_assign=1,
             fast_mode=False, init_population=None, max_budgets=None,
             F=0.5, CR=0.9):
    """
    MODE（差分进化）主循环
    """
    # run_mode 开头
    if pop_size < 4:
        print("⚠️ run_mode 收到 pop_size < 4，自动提升为 4")
        pop_size = 4

    N = len(enc_costs[0])
    T = num_tasks

    print("✅ 正在初始化初始种群(MODE)...")
    if init_population is not None:
        population = init_population
        print("✅ 使用外部传入的初始种群")
    else:
        population = [generate_random_matrix_with_reach(T, N, min_assign, reach)
                      for _ in range(pop_size)]
        print("✅ 默认生成初始种群")

    # --- 新增：极端情况下的初始化补齐（通常不会触发） ---
    if len(population) < pop_size:
        for _ in range(pop_size - len(population)):
            population.append(generate_random_matrix_with_reach(T, N, min_assign, reach))

    # 评估父代
    parents_objs = parallel_evaluate_population(
        population=population,
        enc_costs=enc_costs,
        enc_quals=enc_quals,
        enc_weights=enc_weights,
        pubkey=pubkey, share0=share0, share1=share1, mu=mu, n=n, n_sq=n_sq,
        min_assign=min_assign, fast_mode=fast_mode, max_budgets=max_budgets
    )

    pop_float = [arr for arr in make_float_from_binary(population)]

    for gen in tqdm(range(num_iter), desc="🌱 MODE 进化轮"):
        # 1) 变异+交叉生成子代（浮点）→ 二值化+修复
        offsprings = []
        for i in range(pop_size):
            u_float = de_mutation_and_crossover(i, pop_float, F=F, CR=CR)
            x_child = binarize_and_repair(u_float, reach, min_assign)
            offsprings.append(x_child)

        # 2) 评估子代
        off_objs = parallel_evaluate_population(
            population=offsprings,
            enc_costs=enc_costs,
            enc_quals=enc_quals,
            enc_weights=enc_weights,
            pubkey=pubkey, share0=share0, share1=share1, mu=mu, n=n, n_sq=n_sq,
            min_assign=min_assign, fast_mode=fast_mode, max_budgets=max_budgets
        )

        # 3) 环境选择（父+子 → 下代）
        population, parents_objs = environment_selection_MODE(
            parents=population, parents_objs=parents_objs,
            offsprings=offsprings, off_objs=off_objs,
            pubkey=pubkey, share0=share0, share1=share1, mu=mu, n=n, n_sq=n_sq,
            pop_size=pop_size, alpha_for_crowding=0.5
        )

        # --- 新增：若极端情况下选择后低于 pop_size，则补齐并评估 ---
        if len(population) < pop_size:
            for _ in range(pop_size - len(population)):
                population.append(generate_random_matrix_with_reach(T, N, min_assign, reach))
            # 仅评估新增个体；为了简单可直接整体重评估一次：
            parents_objs = parallel_evaluate_population(
                population=population,
                enc_costs=enc_costs,
                enc_quals=enc_quals,
                enc_weights=enc_weights,
                pubkey=pubkey, share0=share0, share1=share1, mu=mu, n=n, n_sq=n_sq,
                min_assign=min_assign, fast_mode=fast_mode, max_budgets=max_budgets
            )

        # 同步浮点表示（以当前二值为基，回填到±2.0 以稳定后续 DE）
        pop_float = [np.where(np.array(x)>0.5, 2.0, -2.0).astype(float) for x in population]

    print("🎯 正在从最终个体中选择最优解 (Score=α·qual/β·cost)...")
    best_idx = select_best_by_custom_score(
        obj_list=parents_objs,
        share0=share0, share1=share1, pubkey=pubkey, mu=mu, n=n, n_sq=n_sq,
        alpha=0.7, beta=0.3
    )
    best_x, best_cost, best_qual = parents_objs[best_idx]

    print("\n🏆 MODE 优化完成，最优任务分配方案如下：")
    for t in range(T):
        print(f"  任务{t + 1}: {best_x[t]}")

    return best_x, best_cost, best_qual, parents_objs


import os, json
from datetime import datetime

import json
import time
from datetime import datetime
from openpyxl import load_workbook

# === Excel 写入安全转换，避免 OverflowError / NaN/Inf ===
EXCEL_SAFE_INT = 2**53 - 1  # Excel(IEEE-754 double) 约 15 位十进制精度

try:
    import numpy as _np
except Exception:
    _np = None

def _is_np_int(x):
    return _np is not None and isinstance(x, _np.integer)

def _is_np_float(x):
    return _np is not None and isinstance(x, _np.floating)

def coerce_for_excel(v):
    # numpy 标量 -> Python 原生
    if _is_np_int(v) or _is_np_float(v):
        v = v.item()

    # 特大整数 -> 转成字符串，避免 openpyxl 尝试 float() 溢出
    if isinstance(v, int):
        return v if abs(v) <= EXCEL_SAFE_INT else str(v)

    # 浮点 NaN/Inf -> 字符串（Excel 不接受）
    if isinstance(v, float):
        if not math.isfinite(v):
            return str(v)
        return v

    # 其它对象（包括自定义对象）保持原样或字符串化
    return v

def append_row_safe(ws, row):
    ws.append([coerce_for_excel(x) for x in row])

def save_pareto_set_to_excel(xlsx_path: str, params: dict, pareto_set: list,
                             runtime_sec: float, pubkey, share0, share1, mu, n, n_sq):
    """将非支配解集保存到 Excel 文件"""
    xlsx_path = ensure_results_sheet(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb["Results"]

    for sol in pareto_set:
        x, enc_cost, enc_qual = sol

        # 解密
        dec_cost = combine_shares(
            partial_decrypt(enc_cost, share0, pubkey, n_sq),
            partial_decrypt(enc_cost, share1, pubkey, n_sq),
            mu, n
        )
        dec_qual = combine_shares(
            partial_decrypt(enc_qual, share0, pubkey, n_sq),
            partial_decrypt(enc_qual, share1, pubkey, n_sq),
            mu, n
        )
        # 映射回有符号
        dec_cost = to_signed_mod(dec_cost, n)
        dec_qual = to_signed_mod(dec_qual, n)

        # ✅ 用 append_row_safe，自动把超大整数/NaN/Inf 转为字符串
        append_row_safe(ws, [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            int(params["num_tasks"]), int(params["num_workers"]), int(params["min_assign"]),
            int(params["num_iter"]), int(params["pop_size"]),
            int(dec_cost), int(dec_qual), len(pareto_set),
            round(runtime_sec, 2),  # 直接 round，避免先转字符串再转回 float
            json.dumps(x, ensure_ascii=False)
        ])

    # ✅ 保存并显式关闭，减少 Windows 上临时文件占用
    wb.save(xlsx_path)
    wb.close()


def ensure_results_sheet(xlsx_path: str):
    from openpyxl import Workbook, load_workbook

    dir_ = os.path.dirname(os.path.abspath(xlsx_path))
    if dir_ and not os.path.exists(dir_):
        os.makedirs(dir_, exist_ok=True)

    header = [
        "timestamp","num_tasks","num_workers","min_assign","num_iter","pop_size",
        "best_cost","best_qual","pareto_count","runtime_sec","best_x_json"
    ]

    if not os.path.exists(xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(header)
        wb.save(xlsx_path)
        wb.close()
        return xlsx_path

    wb = load_workbook(xlsx_path)
    if "Results" not in wb.sheetnames:
        ws = wb.create_sheet("Results")
        ws.append(header)
        wb.save(xlsx_path)
    wb.close()
    return xlsx_path


def append_run_result(xlsx_path: str, params: dict, best_cost: int, best_qual: int,
                      pareto_count: int, runtime_sec: float, best_x):
    """将一次运行结果追加到 Results 表"""
    xlsx_path = ensure_results_sheet(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb["Results"]

    append_row_safe(ws, [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        int(params["num_tasks"]), int(params["num_workers"]), int(params["min_assign"]),
        int(params["num_iter"]), int(params["pop_size"]),
        int(best_cost), int(best_qual), int(pareto_count),
        round(runtime_sec, 2),
        json.dumps(best_x, ensure_ascii=False)
    ])

    wb.save(xlsx_path)
    wb.close()

# -----------------------------
# 程序入口（调用 MODE 版本）
# -----------------------------
def main():

    # ✅ 参数设置（从 Excel 读取多组；每行一组参数）
    import pandas as pd

    excel_path = 'data_mode.xlsx'  # 你的参数表；Sheet 名：Config
    df_cfg = pd.read_excel(excel_path, sheet_name='Config')

    # 要求表头包含以下列名（不区分大小写也可自己统一转小写）：
    # num_tasks, num_workers, min_assign, num_iter, pop_size
    need = {'num_tasks', 'num_workers', 'min_assign', 'num_iter', 'pop_size'}
    lowermap = {c.lower(): c for c in df_cfg.columns}
    if not need.issubset(lowermap.keys()):
        raise ValueError(f"Config表头缺少列：{need - set(lowermap.keys())}，请确保含有 {need}")

    # 循环每一行参数组合：对每一组都完整执行下面的原始流程
    for _row_i, row in df_cfg.iterrows():

        print(f"\n==================== 参数组合 #{_row_i + 1} ====================")
        num_tasks = int(row[lowermap['num_tasks']])
        num_workers = int(row[lowermap['num_workers']])
        min_assign = int(row[lowermap['min_assign']])
        num_iter = int(row[lowermap['num_iter']])
        pop_size = int(row[lowermap['pop_size']])

        print(f"本次运行参数：num_tasks={num_tasks}, num_workers={num_workers}, "
              f"min_assign={min_assign}, num_iter={num_iter}, pop_size={pop_size}")

        # --- 新增：DE 至少需要 4 个个体 ---
        if pop_size < 4:
            print("⚠️ Config 中的 pop_size < 4，已自动提升为 4（DE 至少需要 4 个个体）")
            pop_size = 4

        # 执行每组参数 5 次
        for run in range(10):  # 内部循环，执行 5 次
            print(f"执行第 {run + 1} 次...")
            # 清空旧文件
            for file in ["enc_worker_data.pkl", "threshold_key_shares.pkl", "offline_cache.pkl"]:
                if os.path.exists(file):
                    os.remove(file)

            print("🔐 加密众包优化系统（MODE 对比版）启动...")
            # 上传加密数据（含位置与最大距离）
            simulate_worker_upload(num_tasks=num_tasks, num_workers=num_workers)

            (pubkey, enc_costs, enc_quals, enc_weights,
             enc_task_locs, enc_worker_locs, enc_max_dists,
             raw_costs, raw_quals, raw_weights,
             raw_task_locs, raw_worker_locs, raw_max_dists,
             share0, share1, mu, n, n_sq, max_budgets) = prepare_data(num_tasks=num_tasks, num_workers=num_workers)

            print("\n📌 工人成本权重向量:")
            print(raw_weights)

            print("\n📌 工人成本矩阵:")
            for i in range(num_workers):
                worker_costs = [raw_costs[t][i] for t in range(num_tasks)]
                print(f"工人 {i + 1}: {worker_costs}")

            print("\n📌 工人位置与最大可接受距离:")
            for i, ((x, y), dmax) in enumerate(zip(raw_worker_locs, raw_max_dists)):
                print(f"工人 {i + 1}: ({x}, {y})，最大距离: {dmax}")

            print("\n📌 任务质量矩阵:")
            for t in range(num_tasks):
                print(f"任务 {t+1}: {raw_quals[t]}")

            print("\n📌 任务位置坐标:")
            for t, (x, y) in enumerate(raw_task_locs):
                print(f"任务 {t + 1}: ({x}, {y})")

            # 离线缓存
            if not os.path.exists("offline_cache.pkl"):
                generate_offline_cache(pubkey, num_sets=300, num_workers=16)
            print(f"✅ 已加载离线缓存")

            # reach[t][i]
            print("\n🧮 正在计算可达性矩阵 reach[t][i]...")
            reach = check_reachability_matrix(enc_task_locs, enc_worker_locs, enc_max_dists,
                                              pubkey, share0, share1, mu, n, n_sq)
            print("\n📌 可达性矩阵（1=可分配，0=超距）:")
            for t in range(num_tasks):
                print(f"任务 {t+1}: {reach[t]}")
            check_task_reachability(reach)

            # 初始化种群（考虑可达性）
            population = [generate_random_matrix_with_reach(num_tasks, num_workers, min_assign, reach)
                          for _ in range(pop_size)]

            # 启动 MODE 优化
            start_time = time.time()
            best_x, enc_best_cost, enc_best_qual, obj_list = run_mode(
                pubkey=pubkey,
                enc_costs=enc_costs,
                enc_quals=enc_quals,
                enc_weights=enc_weights,
                share0=share0, share1=share1, mu=mu, n=n, n_sq=n_sq,
                reach=reach,
                num_iter=num_iter, pop_size=pop_size,
                num_tasks=num_tasks, min_assign=min_assign,
                fast_mode=False,
                init_population=population,
                max_budgets=max_budgets,
                F=0.5, CR=0.9
            )

            print("当前 obj_list（最终一代）内容：")
            for i, (x, enc_cost, enc_qual) in enumerate(obj_list):
                dec_cost = combine_shares(
                    partial_decrypt(enc_cost, share0, pubkey, n_sq),
                    partial_decrypt(enc_cost, share1, pubkey, n_sq),
                    mu, n
                )
                dec_qual = combine_shares(
                    partial_decrypt(enc_qual, share0, pubkey, n_sq),
                    partial_decrypt(enc_qual, share1, pubkey, n_sq),
                    mu, n
                )
                dec_cost = to_signed_mod(dec_cost, n)
                dec_qual = to_signed_mod(dec_qual, n)
                print(f"解 {i}: 成本 = {dec_cost}, 质量 = {dec_qual}")

            pareto_indices = fast_non_dominated_sort_enc(obj_list, pubkey, share0, share1, mu, n, n_sq)[0]
            print(f"\n🎯 非支配解集个数：{len(pareto_indices)}")
            for i in pareto_indices:
                x, enc_cost, enc_qual = obj_list[i]
                dec_cost = combine_shares(
                    partial_decrypt(enc_cost, share0, pubkey, n_sq),
                    partial_decrypt(enc_cost, share1, pubkey, n_sq),
                    mu, n
                )
                dec_qual = combine_shares(
                    partial_decrypt(enc_qual, share0, pubkey, n_sq),
                    partial_decrypt(enc_qual, share1, pubkey, n_sq),
                    mu, n
                )
                print(f"🧬 解 {i}: 成本 = {dec_cost}, 质量 = {dec_qual}")

            print(f"\n✅ 最终任务分配方案：")
            for t, task_assign in enumerate(best_x):
                print(f"任务 {t + 1}: {task_assign}")

            dec_best_cost = combine_shares(
                partial_decrypt(enc_best_cost, share0, pubkey, n_sq),
                partial_decrypt(enc_best_cost, share1, pubkey, n_sq),
                mu, n
            )
            dec_best_qual = combine_shares(
                partial_decrypt(enc_best_qual, share0, pubkey, n_sq),
                partial_decrypt(enc_best_qual, share1, pubkey, n_sq),
                mu, n
            )
            print(f"\n🔓 解密后目标值：")
            print(f"📉 总成本 = {dec_best_cost}")
            print(f"📈 总质量 = {dec_best_qual}")
            print(f"\n⏱️ 本次运行耗时：{time.time() - start_time:.2f} 秒")
            # ===== 你的原始逻辑到此结束；下一轮循环自动开始 =====

            # ===== 写入 Results 工作表 =====
            excel_path = "data_mode.xlsx"  # 自己的结果文件名/路径
            params_for_log = {
                "num_tasks": num_tasks,
                "num_workers": num_workers,
                "min_assign": min_assign,
                "num_iter": num_iter,
                "pop_size": pop_size
            }
            runtime = time.time() - start_time
            pareto_indices = fast_non_dominated_sort_enc(obj_list, pubkey, share0, share1, mu, n, n_sq)[0]

            # 保存所有非支配解集到 Excel
            print(f"✅ 正在保存非支配解集到 Excel ...")
            save_pareto_set_to_excel(
                xlsx_path=excel_path,
                params=params_for_log,
                pareto_set=[obj_list[i] for i in pareto_indices],
                runtime_sec=runtime,
                pubkey=pubkey,
                share0=share0,
                share1=share1,
                mu=mu,
                n=n,
                n_sq=n_sq
            )

            print(f"✅ 结果已保存到 {excel_path} -> Results")

if __name__ == "__main__":
    main()
